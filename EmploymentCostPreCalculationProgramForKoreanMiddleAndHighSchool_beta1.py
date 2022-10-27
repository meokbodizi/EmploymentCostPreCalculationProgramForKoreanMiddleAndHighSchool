# -*- coding: utf-8 -*-
"""
Created on Sun Jan  9 19:33:26 2022

@author: JinseongKim
"""

from tkinter import *
import tkinter.ttk as ttk
from tkinter.messagebox import showinfo
import re
from datetime import datetime
from dateutil.relativedelta import relativedelta
from urllib.request import urlopen
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import pickle
from functools import reduce
import json
import numpy as np

if __name__ == "__main__":

    employeelist = []

    root = Tk()
    root.title("HRSystem")
    root.geometry("800x640")

    # menu = Menu(root)
    # menu_file = Menu(menu, tearoff=0)
    # menu_file.add_command(label="백업", command=lambda :None)
    # menu_file.add_command(label="저장", command=lambda :None)
    # menu_file.add_separator()
    # menu_file.add_command(label="종료", command=root.destroy)
    # menu.add_cascade(label="파일", menu=menu_file)

    employeelist_label = Label(root, text="교직원목록")
    employeelist_label.place(x=0, y=0)

    employeelist_treeview = ttk.Treeview(root, height = 15, columns=('id','성명'), show="headings")
    employeelist_treeview.heading("id",text="id")
    employeelist_treeview.heading("성명",text="성명")
    [employeelist_treeview.insert("",END,values=(employee["id"],employee["성명"])) for employee in employeelist]
    employeelist_treeview.column(0, width=25)
    employeelist_treeview.column(1, width=50)
    employeelist_treeview.place(x=0, y=20)

    def employee_search_name_entry_chk(input_name):
        if re.search("^[가-힣]{1,4}$", input_name):
            return True
        elif input_name =="":
            return True
        else:
            return False
    def search_employee_name(*args):
        for item in employeelist_treeview.get_children():
            if employeelist_treeview.item(item)["values"][1] == employee_search_name_entry_var.get():
                employeelist_treeview.selection_set(item)
                # employeelist_treeview.yview(30)
    employee_search_name_entry_chkreg=root.register(employee_search_name_entry_chk)
    employee_search_name_label = Label(root, text="검색창")
    employee_search_name_label.place(x=0, y=350)
    employee_search_name_entry_var = StringVar()
    employee_search_name_entry_var.trace("w", search_employee_name)
    employee_search_name_entry = Entry(root, width=6, textvariable=employee_search_name_entry_var)
    employee_search_name_entry.config(validate="key", validatecommand=(employee_search_name_entry_chkreg,"%P"))
    employee_search_name_entry.place(x=50, y=350)


    def enable_grade():
        employee_gradebtn.config(state="readonly")

    def disable_grade():
        employee_gradebtn.current(0)
        employee_gradebtn.config(state="disabled")

    def on_selection_treeview(event):
        selection = event.widget.selection()
        if selection:
            item =  event.widget.item(selection[0])
            record = item['values']
            data = [employee for employee in employeelist if employee["id"] == record[0]][0]
            employee_id_entry.delete(0,END)
            employee_id_entry.insert(0,data["id"])
            employee_name_entry.delete(0,END)
            employee_name_entry.insert(0,data["성명"])
            employee_rrn1_entry.delete(0,END)
            employee_rrn1_entry.insert(0,data["주민번호"].split('-')[0])
            employee_rrn2_entry.delete(0,END)
            employee_rrn2_entry.insert(0,data["주민번호"].split('-')[1])
            if data["직종"]=="행정직":
                employee_category_radiobtn1.select()
                enable_grade()
                employee_anual_leave_compensation_entry.config(state="normal")
                employee_anual_leave_compensation_entry.delete(0,END)
                if "연가보상일수" in data.keys():
                    employee_anual_leave_compensation_entry.insert(0,data["연가보상일수"])
            elif data["직종"]=="교원":
                employee_category_radiobtn2.select()
                disable_grade()
                employee_anual_leave_compensation_entry.delete(0,END)
                employee_anual_leave_compensation_entry.config(state="disabled")
            elif data["직종"]=="기간제교원":
                employee_category_radiobtn3.select()
                disable_grade()
                employee_anual_leave_compensation_entry.delete(0,END)
                employee_anual_leave_compensation_entry.config(state="disabled")
            employee_year_pay_increasebtn.current(employee_year_pay_increasebtn_var.index(data["승급년월일"].split("-")[0]+"년"))
            employee_month_pay_increasebtn.current(employee_month_pay_increasebtn_var.index(data["승급년월일"].split("-")[1]+"월"))
            employee_day_pay_increasebtn.current(employee_day_pay_increasebtn_var.index(data["승급년월일"].split("-")[2]+"일"))
            employee_year_datechangebtn.current(employee_year_datechangebtn_var.index(data["현근무년수변경일"].split("-")[0]+"년"))
            employee_month_datechangebtn.current(employee_month_datechangebtn_var.index(data["현근무년수변경일"].split("-")[1]+"월"))
            employee_day_datechangebtn.current(employee_day_datechangebtn_var.index(data["현근무년수변경일"].split("-")[2]+"일"))
            employee_years_of_service_year_entry.delete(0,END)
            employee_years_of_service_year_entry.insert(0,data["근무연한"][0])
            employee_years_of_service_month_entry.delete(0,END)
            employee_years_of_service_month_entry.insert(0,data["근무연한"][1])
            employee_years_of_service_day_entry.delete(0,END)
            employee_years_of_service_day_entry.insert(0,data["근무연한"][2])
            employee_step_entry.delete(0,END)
            employee_step_entry.insert(0,data["호봉"])
            employee_gradebtn.current(employee_gradebtn_var.index(data["급"]))
            if "현부서임용일" in data.keys() and data["현부서임용일"] != "":
                employee_appointment_year_box.current(employee_appointment_year_box_var.index(data["현부서임용일"].split("-")[0]))
                employee_appointment_month_box.current(employee_appointment_month_box_var.index(data["현부서임용일"].split("-")[1]))
                employee_appointment_day_box.current(employee_appointment_day_box_var.index(data["현부서임용일"].split("-")[2]))
            else:
                employee_appointment_year_box.current(0)
                employee_appointment_month_box.current(0)
                employee_appointment_day_box.current(0)
            if "계속근무여부" in data.keys() and data["계속근무여부"]==1:
                employee_keep_working_radiobtn2.select()
            else:
                employee_keep_working_radiobtn1.select()
            if "퇴직일" in data.keys() and data["퇴직일"] != "":
                employee_retire_year_box.current(employee_retire_year_box_var.index(data["퇴직일"].split("-")[0]))
                employee_retire_month_box.current(employee_retire_month_box_var.index(data["퇴직일"].split("-")[1]))
                employee_retire_day_box.current(employee_retire_day_box_var.index(data["퇴직일"].split("-")[2]))
            else:
                employee_retire_year_box.current(0)
                employee_retire_month_box.current(0)
                employee_retire_day_box.current(0)
            employee_position_btn.current(employee_position_btn_var.index(data["보직"]))
            employee_special_class_btn.current(employee_special_class_btn_var.index(data["가산정원"]))
            if "원로교사" in data.keys() and data["원로교사"]==1:
                employee_elder_radiobtn2.select()
            else:
                employee_elder_radiobtn1.select()
            for item in employee_absence_treeview.get_children():
                employee_absence_treeview.delete(item)
            if "휴직" in data.keys():
                [employee_absence_treeview.insert("",END,values=(category, datetime_start, datetime_end, contigous)) for category, datetime_start, datetime_end, contigous in data["휴직"]]
            for item in employee_career_treeview.get_children():
                employee_career_treeview.delete(item)
            if "경력" in data.keys():
                [employee_career_treeview.insert("",END,values=(category, position, public, rate, datetime_start, datetime_end)) for category, position, public, rate, datetime_start, datetime_end in data["경력"]]
            if "감봉시작일" in data.keys() and data["감봉시작일"] != "":
                employee_paycut_start_year_box.current(employee_paycut_start_year_box_var.index(data["감봉시작일"].split("-")[0]))
                employee_paycut_start_month_box.current(employee_paycut_start_month_box_var.index(data["감봉시작일"].split("-")[1]))
                employee_paycut_start_day_box.current(employee_paycut_start_day_box_var.index(data["감봉시작일"].split("-")[2]))
            else:
                employee_paycut_start_year_box.current(0)
                employee_paycut_start_month_box.current(0)
                employee_paycut_start_day_box.current(0)
            if "감봉종료일" in data.keys() and data["감봉종료일"] != "":
                employee_paycut_end_year_box.current(employee_paycut_end_year_box_var.index(data["감봉종료일"].split("-")[0]))
                employee_paycut_end_month_box.current(employee_paycut_end_month_box_var.index(data["감봉종료일"].split("-")[1]))
                employee_paycut_end_day_box.current(employee_paycut_end_day_box_var.index(data["감봉종료일"].split("-")[2]))
            else:
                employee_paycut_end_year_box.current(0)
                employee_paycut_end_month_box.current(0)
                employee_paycut_end_day_box.current(0)
            employee_paycut_rate_entry.delete(0,END)
            if "감봉율" in data.keys():
                employee_paycut_rate_entry.insert(0,data["감봉율"])
            if "승급제한시작일" in data.keys() and data["승급제한시작일"] != "":
                employee_upgrade_restriction_start_year_box.current(employee_upgrade_restriction_start_year_box_var.index(data["승급제한시작일"].split("-")[0]))
                employee_upgrade_restriction_start_month_box.current(employee_upgrade_restriction_start_month_box_var.index(data["승급제한시작일"].split("-")[1]))
                employee_upgrade_restriction_start_day_box.current(employee_upgrade_restriction_start_day_box_var.index(data["승급제한시작일"].split("-")[2]))
            else:
                employee_upgrade_restriction_start_year_box.current(0)
                employee_upgrade_restriction_start_month_box.current(0)
                employee_upgrade_restriction_start_day_box.current(0)
            if "승급제한종료일" in data.keys() and data["승급제한종료일"] != "":
                employee_upgrade_restriction_end_year_box.current(employee_upgrade_restriction_end_year_box_var.index(data["승급제한종료일"].split("-")[0]))
                employee_upgrade_restriction_end_month_box.current(employee_upgrade_restriction_end_month_box_var.index(data["승급제한종료일"].split("-")[1]))
                employee_upgrade_restriction_end_day_box.current(employee_upgrade_restriction_end_day_box_var.index(data["승급제한종료일"].split("-")[2]))
            else:
                employee_upgrade_restriction_end_year_box.current(0)
                employee_upgrade_restriction_end_month_box.current(0)
                employee_upgrade_restriction_end_day_box.current(0)
    employeelist_treeview.bind("<<TreeviewSelect>>", on_selection_treeview)

    # employeelist_treeview_scrollbar = ttk.Scrollbar(root, orient=VERTICAL, command=employeelist_treeview.yview)
    # employeelist_treeview.configure(yscroll=employeelist_treeview_scrollbar.set)
    # employeelist_treeview_scrollbar.place(x=70,y=0)

    def epmployee_entry_number_register(maxlength):
        def epmployee_entry_number_chk(input_string):
            if input_string.isdigit() and len(input_string)<=maxlength:
                return True
            elif input_string =="":
                return True
            else:
                return False
        return root.register(epmployee_entry_number_chk)

    employee_id_label = Label(root, text="ID")
    employee_id_label.place(x=80+100, y=0)
    employee_id_entry = Entry(root, width=2)
    employee_id_entry.config(validate="key", validatecommand=(epmployee_entry_number_register(999),"%P"))
    employee_id_entry.place(x=100+100, y=0)

    def entry_limit_number_quantity_length_register(maxlength, maxnumber):
        def entry_limit_number_quantity_length_chk(input_string):
            if input_string.isdigit() and len(input_string)<=maxlength and int(input_string)<=maxnumber:
                return True
            elif input_string =="":
                return True
            else:
                return False
        return root.register(entry_limit_number_quantity_length_chk)

    def employee_name_entry_chk(input_name):
        if re.search("[가-힣]{1,4}", input_name):
            return True
        elif input_name =="":
            return True
        else:
            return False
    employee_name_entry_chkreg=root.register(employee_name_entry_chk)

    employee_name_label = Label(root, text="성명")
    employee_name_label.place(x=120+100, y=0)
    employee_name_entry = Entry(root, width=5)
    employee_name_entry.config(validate="key", validatecommand=(employee_name_entry_chkreg,"%P"))
    employee_name_entry.place(x=150+100, y=0)


    employee_rrn1_label = Label(root, text="주민등록번호")
    employee_rrn1_label.place(x=80+100, y=25)
    employee_rrn1_entry = Entry(root, width=7)
    employee_rrn1_entry.config(validate="key", validatecommand=(epmployee_entry_number_register(6),"%P"))
    employee_rrn1_entry.place(x=80+100, y=50)
    employee_rrn2_label = Label(root, text="-")
    employee_rrn2_label.place(x=130+100, y=50)
    employee_rrn2_entry = Entry(root, width=8)
    employee_rrn2_entry.config(validate="key", validatecommand=(epmployee_entry_number_register(7),"%P"))
    employee_rrn2_entry.place(x=145+100, y=50)

    def enable_grade_and_anual_leave_compensation():
        enable_grade()
        employee_anual_leave_compensation_entry.config(state="normal")
    def disable_grade_and_anual_leave_compensation():
        disable_grade()
        employee_anual_leave_compensation_entry.delete(0,END)
        employee_anual_leave_compensation_entry.config(state="disabled")

    employee_category_radiobtn_var = StringVar()
    employee_category_radiobtn1 = Radiobutton(root, text="행정직", value="행정직", variable=employee_category_radiobtn_var, command=enable_grade_and_anual_leave_compensation)
    employee_category_radiobtn1.select()
    employee_category_radiobtn2 = Radiobutton(root, text="교원", value="교원", variable=employee_category_radiobtn_var, command=disable_grade_and_anual_leave_compensation)
    employee_category_radiobtn3 = Radiobutton(root, text="기간제교원", value="기간제교원", variable=employee_category_radiobtn_var, command=disable_grade_and_anual_leave_compensation)
    employee_category_radiobtn1.place(x=80, y=0)
    employee_category_radiobtn2.place(x=80, y=25)
    employee_category_radiobtn3.place(x=80, y=50)


    employee_date_pay_increasement_label = Label(root, text='승급년월일')
    employee_date_pay_increasement_label.place(x=310-45, y=25)

    employee_year_pay_increasebtn_var = [str(i)+'년' for i in range(datetime.now().year,datetime.now().year-5,-1)]
    employee_year_pay_increasebtn = ttk.Combobox(root, height=5, width=6, state='readonly', values=employee_year_pay_increasebtn_var)
    employee_year_pay_increasebtn.current(0)
    employee_year_pay_increasebtn.place(x=375-45, y=25)
    employee_month_pay_increasebtn_var = [str(i)+'월' for i in range(1,13)]
    employee_month_pay_increasebtn = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_month_pay_increasebtn_var)
    employee_month_pay_increasebtn.current(0)
    employee_month_pay_increasebtn.place(x=445-45, y=25)
    def update_employee_pay_increase_day_box(event):
        if employee_year_pay_increasebtn.get() == "":
            showinfo("경고", "연도를 선택해주십시오.")
            return None
        employee_day_pay_increasebtn['values'] = [""] + [str(i)+'일' for i in range(1, (datetime(int(employee_year_pay_increasebtn.get().replace("년","")), int(employee_month_pay_increasebtn.get().replace("월","")), 1) + relativedelta(months=1) - relativedelta(days=1)).day+1)]
        employee_day_pay_increasebtn.current(0)
    employee_month_pay_increasebtn.bind('<<ComboboxSelected>>', update_employee_pay_increase_day_box)
    employee_day_pay_increasebtn_var = [str(i)+'일' for i in range(1,31)]
    employee_day_pay_increasebtn = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_day_pay_increasebtn_var)
    employee_day_pay_increasebtn.current(0)
    employee_day_pay_increasebtn.place(x=500-45, y=25)


    employee_date_datechange_label = Label(root, text='현근무연수변경일')
    employee_date_datechange_label.place(x=310+200, y=25)

    employee_year_datechangebtn_var = [str(i)+'년' for i in range(datetime.now().year,datetime.now().year-5,-1)]
    employee_year_datechangebtn = ttk.Combobox(root, height=5, width=6, state='readonly', values=employee_year_datechangebtn_var)
    employee_year_datechangebtn.current(0)
    employee_year_datechangebtn.place(x=375+200+40, y=25)
    employee_month_datechangebtn_var = [str(i)+'월' for i in range(1,13)]
    employee_month_datechangebtn = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_month_datechangebtn_var)
    employee_month_datechangebtn.current(0)
    employee_month_datechangebtn.place(x=445+200+40, y=25)
    def update_employee_datechange_day_box(event):
        if employee_year_datechangebtn.get() == "":
            showinfo("경고", "연도를 선택해주십시오.")
            return None
        employee_day_datechangebtn['values'] = [""] + [str(i)+'일' for i in range(1, (datetime(int(employee_year_datechangebtn.get().replace("년","")), int(employee_month_datechangebtn.get().replace("월","")), 1) + relativedelta(months=1) - relativedelta(days=1)).day+1)]
        employee_day_datechangebtn.current(0)
    employee_month_datechangebtn.bind('<<ComboboxSelected>>', update_employee_datechange_day_box)
    employee_day_datechangebtn_var = [str(i)+'일' for i in range(1,32)]
    employee_day_datechangebtn = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_day_datechangebtn_var)
    employee_day_datechangebtn.current(0)
    employee_day_datechangebtn.place(x=500+200+40, y=25)


    employee_years_of_service_label = Label(root, text="근무연한")
    employee_years_of_service_label.place(x=310+230, y=0)
    employee_years_of_service_year_entry = Entry(root, width=2)
    employee_years_of_service_year_entry.config(validate="key", validatecommand=(epmployee_entry_number_register(2),"%P"))
    employee_years_of_service_year_entry.place(x=370+230, y=0)
    employee_years_of_service_year_label = Label(root, text="년")
    employee_years_of_service_year_label.place(x=385+230, y=0)
    employee_years_of_service_month_entry = Entry(root, width=2)
    employee_years_of_service_month_entry.config(validate="key", validatecommand=(entry_limit_number_quantity_length_register(2,12),"%P"))
    employee_years_of_service_month_entry.place(x=370+230+40, y=0)
    employee_years_of_service_month_label = Label(root, text="월")
    employee_years_of_service_month_label.place(x=385+230+40, y=0)
    employee_years_of_service_day_entry = Entry(root, width=2)
    employee_years_of_service_day_entry.config(validate="key", validatecommand=(entry_limit_number_quantity_length_register(2,31),"%P"))
    employee_years_of_service_day_entry.place(x=370+230+80, y=0)
    employee_years_of_service_day_label = Label(root, text="일")
    employee_years_of_service_day_label.place(x=385+230+80, y=0)

    employee_step_label = Label(root, text="호봉")
    employee_step_label.place(x=425-90, y=0)
    employee_step_entry = Entry(root, width=2)
    employee_step_entry.config(validate="key", validatecommand=(epmployee_entry_number_register(2),"%P"))
    employee_step_entry.place(x=410-90, y=0)

    employee_gradebtn_var = [""] + [str(i)+'급' for i in range(1,10)]
    employee_gradebtn= ttk.Combobox(root, height=5, width=3, state='readonly', values=employee_gradebtn_var)
    employee_gradebtn.current(0)
    employee_gradebtn.place(x=460-90, y=0)


    employee_anual_leave_compensation_label = Label(root, text="연가보상일수")
    employee_anual_leave_compensation_label.place(x=510-90, y=0)
    employee_anual_leave_compensation_entry = Entry(root, width=2)
    employee_anual_leave_compensation_entry.config(validate="key", validatecommand=(entry_limit_number_quantity_length_register(2, 20),"%P"))
    employee_anual_leave_compensation_entry.place(x=590-90, y=0)

    employee_position_label = Label(root, text='보직')
    employee_position_label.place(x=310, y=50)
    employee_position_btn_var = ["","교장","교감","부장담임","부장","담임","교사","행정실장","주무관"]
    employee_position_btn = ttk.Combobox(root, height=5, width=8, state='readonly', values=employee_position_btn_var)
    employee_position_btn.current(0)
    employee_position_btn.place(x=340, y=50)

    employee_special_class_label = Label(root, text='가산정원')
    employee_special_class_label.place(x=420, y=50)
    employee_special_class_btn_var = ["","보건","상담","사서","영양"]
    employee_special_class_btn = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_special_class_btn_var)
    employee_special_class_btn.current(0)
    employee_special_class_btn.place(x=475, y=50)


    employee_elder_label = Label(root, text='원로교사')
    employee_elder_label.place(x=530, y=50)
    employee_elder_radiobtn_var = IntVar()
    employee_elder_radiobtn1 = Radiobutton(root, text="N", value=0, variable=employee_elder_radiobtn_var)
    employee_elder_radiobtn1.place(x=580,y=50)
    employee_elder_radiobtn1.select()
    employee_elder_radiobtn2 = Radiobutton(root, text="Y", value=1, variable=employee_elder_radiobtn_var)
    employee_elder_radiobtn2.place(x=620,y=50)

    employee_appointment_label = Label(root, text="현부서임용일")
    employee_appointment_label.place(x=510-470+100, y=75)
    employee_appointment_year_label = Label(root, text="년")
    employee_appointment_year_label.place(x=610-440+100, y=75)
    employee_appointment_year_box_var = [""]+[str(i) for i in range(datetime.now().year+1,1950,-1)]
    employee_appointment_year_box = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_appointment_year_box_var)
    employee_appointment_year_box.current(0)
    employee_appointment_year_box.place(x=560-440+100, y=75)
    employee_appointment_month_label = Label(root, text="월")
    employee_appointment_month_label.place(x=665-440+100, y=75)
    employee_appointment_month_box_var = [""]+[f"{i}" for i in range(1,13)]
    employee_appointment_month_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_appointment_month_box_var)
    employee_appointment_month_box.current(0)
    employee_appointment_month_box.place(x=630-440+100, y=75)
    def update_employee_appointment_day_box(event):
        if employee_appointment_year_box.get() == "":
            showinfo("경고", "연도를 선택해주십시오.")
            return None
        employee_appointment_day_box['values'] = [""] + [str(i) for i in range(1, (datetime(int(employee_appointment_year_box.get()), int(employee_appointment_month_box.get()), 1) + relativedelta(months=1) - relativedelta(days=1)).day+1)]
    employee_appointment_month_box.bind('<<ComboboxSelected>>', update_employee_appointment_day_box)
    employee_appointment_day_label = Label(root, text="일")
    employee_appointment_day_label.place(x=720-440+100, y=75)
    employee_appointment_day_box_var = [""]+[str(i) for i in range(1,32)]
    employee_appointment_day_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_appointment_day_box_var)
    employee_appointment_day_box.current(0)
    employee_appointment_day_box.place(x=685-440+100, y=75)

    employee_keep_working_label = Label(root, text='기간제교원 계속근무여부(정근수당 관련)')
    employee_keep_working_label.place(x=140, y=100)
    employee_keep_working_radiobtn_var = IntVar()
    employee_keep_working_radiobtn1 = Radiobutton(root, text="N", value=0, variable=employee_keep_working_radiobtn_var)
    employee_keep_working_radiobtn1.place(x=400-20,y=100)
    employee_keep_working_radiobtn1.select()
    employee_keep_working_radiobtn2 = Radiobutton(root, text="Y", value=1, variable=employee_keep_working_radiobtn_var)
    employee_keep_working_radiobtn2.place(x=440-20,y=100) 

    employee_retire_label = Label(root, text="퇴직일")
    employee_retire_label.place(x=510-200+100, y=75)
    employee_retire_year_label = Label(root, text="년")
    employee_retire_year_label.place(x=610-200+100, y=75)
    employee_retire_year_box_var = [""]+[str(i) for i in range(datetime.now().year+1,1950,-1)]
    employee_retire_year_box = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_retire_year_box_var)
    employee_retire_year_box.current(0)
    employee_retire_year_box.place(x=560-200+100, y=75)
    employee_retire_month_label = Label(root, text="월")
    employee_retire_month_label.place(x=665-200+100, y=75)
    employee_retire_month_box_var = [""]+[f"{i}" for i in range(1,13)]
    employee_retire_month_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_retire_month_box_var)
    employee_retire_month_box.current(0)
    employee_retire_month_box.place(x=630-200+100, y=75)
    def update_employee_retire_day_box(event):
        if employee_retire_year_box.get() == "":
            showinfo("경고", "연도를 선택해주십시오.")
            return None
        employee_retire_day_box['values'] = [""] + [str(i) for i in range(1, (datetime(int(employee_retire_year_box.get()), int(employee_retire_month_box.get()), 1) + relativedelta(months=1) - relativedelta(days=1)).day+1)]
    employee_retire_month_box.bind('<<ComboboxSelected>>', update_employee_retire_day_box)
    employee_retire_day_label = Label(root, text="일")
    employee_retire_day_label.place(x=720-200+100, y=75)
    employee_retire_day_box_var = [""]+[str(i) for i in range(1,32)]
    employee_retire_day_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_retire_day_box_var)
    employee_retire_day_box.current(0)
    employee_retire_day_box.place(x=685-200+100, y=75)

    employee_paycut_start_label = Label(root, text="감봉기간")
    employee_paycut_start_label.place(x=510-470+100, y=360)
    employee_paycut_start_year_label = Label(root, text="년")
    employee_paycut_start_year_label.place(x=610-440+100, y=360)
    employee_paycut_start_year_box_var = [""]+[str(i) for i in range(datetime.now().year+1,1950,-1)]
    employee_paycut_start_year_box = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_paycut_start_year_box_var)
    employee_paycut_start_year_box.current(0)
    employee_paycut_start_year_box.place(x=560-440+100, y=360)
    employee_paycut_start_month_label = Label(root, text="월")
    employee_paycut_start_month_label.place(x=665-440+100, y=360)
    employee_paycut_start_month_box_var = [""]+[f"{i}" for i in range(1,13)]
    employee_paycut_start_month_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_paycut_start_month_box_var)
    employee_paycut_start_month_box.current(0)
    employee_paycut_start_month_box.place(x=630-440+100, y=360)
    def update_employee_paycut_start_day_box(event):
        if employee_paycut_start_year_box.get() == "":
            showinfo("경고", "연도를 선택해주십시오.")
            return None
        employee_paycut_start_day_box['values'] = [""] + [str(i) for i in range(1, (datetime(int(employee_paycut_start_year_box.get()), int(employee_paycut_start_month_box.get()), 1) + relativedelta(months=1) - relativedelta(days=1)).day+1)]
    employee_paycut_start_month_box.bind('<<ComboboxSelected>>', update_employee_paycut_start_day_box)
    employee_paycut_start_day_label = Label(root, text="일")
    employee_paycut_start_day_label.place(x=720-440+100, y=360)
    employee_paycut_start_day_box_var = [""]+[str(i) for i in range(1,32)]
    employee_paycut_start_day_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_paycut_start_day_box_var)
    employee_paycut_start_day_box.current(0)
    employee_paycut_start_day_box.place(x=685-440+100, y=360)

    employee_paycut_end_label = Label(root, text="~")
    employee_paycut_end_label.place(x=510-470+360, y=360)
    employee_paycut_end_year_label = Label(root, text="년")
    employee_paycut_end_year_label.place(x=610-440+300, y=360)
    employee_paycut_end_year_box_var = [""]+[str(i) for i in range(datetime.now().year+5,1950,-1)]
    employee_paycut_end_year_box = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_paycut_end_year_box_var)
    employee_paycut_end_year_box.current(0)
    employee_paycut_end_year_box.place(x=560-440+300, y=360)
    employee_paycut_end_month_label = Label(root, text="월")
    employee_paycut_end_month_label.place(x=665-440+300, y=360)
    employee_paycut_end_month_box_var = [""]+[f"{i}" for i in range(1,13)]
    employee_paycut_end_month_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_paycut_end_month_box_var)
    employee_paycut_end_month_box.current(0)
    employee_paycut_end_month_box.place(x=630-440+300, y=360)
    def update_employee_paycut_end_day_box(event):
        if employee_paycut_end_year_box.get() == "":
            showinfo("경고", "연도를 선택해주십시오.")
            return None
        employee_paycut_end_day_box['values'] = [""] + [str(i) for i in range(1, (datetime(int(employee_paycut_end_year_box.get()), int(employee_paycut_end_month_box.get()), 1) + relativedelta(months=1) - relativedelta(days=1)).day+1)]
    employee_paycut_end_month_box.bind('<<ComboboxSelected>>', update_employee_paycut_end_day_box)
    employee_paycut_end_day_label = Label(root, text="일")
    employee_paycut_end_day_label.place(x=720-440+300, y=360)
    employee_paycut_end_day_box_var = [""]+[str(i) for i in range(1,32)]
    employee_paycut_end_day_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_paycut_end_day_box_var)
    employee_paycut_end_day_box.current(0)
    employee_paycut_end_day_box.place(x=685-440+300, y=360)

    def employee_paycut_rate_entry_chk(input_name):
        if re.search("^[0-9]{1,3}$", input_name) and int(input_name)<=100:
            return True
        elif input_name =="":
            return True
        else:
            return False
    employee_paycut_rate_entry_chkreg=root.register(employee_paycut_rate_entry_chk)
    employee_paycut_rate_entry = Entry(root, width=4)
    employee_paycut_rate_entry.config(validate="key", validatecommand=(employee_paycut_rate_entry_chkreg,"%P"))
    employee_paycut_rate_entry.place(x=720-440+320, y=360)
    employee_paycut_label = Label(root, text="%")
    employee_paycut_label.place(x=720-440+320+30, y=360)
    employee_paycut_label = Label(root, text="감봉율")
    employee_paycut_label.place(x=720-440+320, y=380)    

    employee_upgrade_restriction_start_label = Label(root, text="승급제한기간")
    employee_upgrade_restriction_start_label.place(x=510-470+100, y=380)
    employee_upgrade_restriction_start_year_label = Label(root, text="년")
    employee_upgrade_restriction_start_year_label.place(x=610-440+100, y=380)
    employee_upgrade_restriction_start_year_box_var = [""]+[str(i) for i in range(datetime.now().year+1,1950,-1)]
    employee_upgrade_restriction_start_year_box = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_upgrade_restriction_start_year_box_var)
    employee_upgrade_restriction_start_year_box.current(0)
    employee_upgrade_restriction_start_year_box.place(x=560-440+100, y=380)
    employee_upgrade_restriction_start_month_label = Label(root, text="월")
    employee_upgrade_restriction_start_month_label.place(x=665-440+100, y=380)
    employee_upgrade_restriction_start_month_box_var = [""]+[f"{i}" for i in range(1,13)]
    employee_upgrade_restriction_start_month_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_upgrade_restriction_start_month_box_var)
    employee_upgrade_restriction_start_month_box.current(0)
    employee_upgrade_restriction_start_month_box.place(x=630-440+100, y=380)
    def update_employee_upgrade_restriction_start_day_box(event):
        if employee_upgrade_restriction_start_year_box.get() == "":
            showinfo("경고", "연도를 선택해주십시오.")
            return None
        employee_upgrade_restriction_start_day_box['values'] = [""] + [str(i) for i in range(1, (datetime(int(employee_upgrade_restriction_start_year_box.get()), int(employee_upgrade_restriction_start_month_box.get()), 1) + relativedelta(months=1) - relativedelta(days=1)).day+1)]
    employee_upgrade_restriction_start_month_box.bind('<<ComboboxSelected>>', update_employee_upgrade_restriction_start_day_box)
    employee_upgrade_restriction_start_day_label = Label(root, text="일")
    employee_upgrade_restriction_start_day_label.place(x=720-440+100, y=380)
    employee_upgrade_restriction_start_day_box_var = [""]+[str(i) for i in range(1,32)]
    employee_upgrade_restriction_start_day_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_upgrade_restriction_start_day_box_var)
    employee_upgrade_restriction_start_day_box.current(0)
    employee_upgrade_restriction_start_day_box.place(x=685-440+100, y=380)

    employee_upgrade_restriction_end_label = Label(root, text="~")
    employee_upgrade_restriction_end_label.place(x=510-470+360, y=380)
    employee_upgrade_restriction_end_year_label = Label(root, text="년")
    employee_upgrade_restriction_end_year_label.place(x=610-440+300, y=380)
    employee_upgrade_restriction_end_year_box_var = [""]+[str(i) for i in range(datetime.now().year+5,1950,-1)]
    employee_upgrade_restriction_end_year_box = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_upgrade_restriction_end_year_box_var)
    employee_upgrade_restriction_end_year_box.current(0)
    employee_upgrade_restriction_end_year_box.place(x=560-440+300, y=380)
    employee_upgrade_restriction_end_month_label = Label(root, text="월")
    employee_upgrade_restriction_end_month_label.place(x=665-440+300, y=380)
    employee_upgrade_restriction_end_month_box_var = [""]+[f"{i}" for i in range(1,13)]
    employee_upgrade_restriction_end_month_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_upgrade_restriction_end_month_box_var)
    employee_upgrade_restriction_end_month_box.current(0)
    employee_upgrade_restriction_end_month_box.place(x=630-440+300, y=380)
    def update_employee_upgrade_restriction_end_day_box(event):
        if employee_upgrade_restriction_end_year_box.get() == "":
            showinfo("경고", "연도를 선택해주십시오.")
            return None
        employee_upgrade_restriction_end_day_box['values'] = [""] + [str(i) for i in range(1, (datetime(int(employee_upgrade_restriction_end_year_box.get()), int(employee_upgrade_restriction_end_month_box.get()), 1) + relativedelta(months=1) - relativedelta(days=1)).day+1)]
    employee_upgrade_restriction_end_month_box.bind('<<ComboboxSelected>>', update_employee_upgrade_restriction_end_day_box)
    employee_upgrade_restriction_end_day_label = Label(root, text="일")
    employee_upgrade_restriction_end_day_label.place(x=720-440+300, y=380)
    employee_upgrade_restriction_end_day_box_var = [""]+[str(i) for i in range(1,32)]
    employee_upgrade_restriction_end_day_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_upgrade_restriction_end_day_box_var)
    employee_upgrade_restriction_end_day_box.current(0)
    employee_upgrade_restriction_end_day_box.place(x=685-440+300, y=380)


    def employee_family_add(root):
        def epmployee_entry_number_register(maxlength):
            def epmployee_entry_number_chk(input_string):
                if input_string.isdigit() and len(input_string)<=maxlength:
                    return True
                elif input_string =="":
                    return True
                else:
                    return False
            return root.register(epmployee_entry_number_chk)
        if employeelist_treeview.selection():
            subroot = Toplevel(root)
            subroot.title("가족사항")
            subroot.geometry("600x300")
            Label(subroot, text="가족관계").place(x=0,y=0)
            ##
            subroot_box = []
            employee_rrn1_entry = []
            employee_rrn2_entry = []
            subroot_box_var = []
            employee_YN_label = []
            employee_YN_radiobtn_var = []
            employee_YN_radiobtn1 = []
            employee_YN_radiobtn2 = []
            subroot_box_var.append(["","배우자","직계존속","직계비속"])
            subroot_box.append(ttk.Combobox(subroot, width=7, state='readonly', values=subroot_box_var[0]))
            subroot_box[0].current(0)
            subroot_box[0].place(x=0,y=25)
            employee_id_label1 = Label(subroot, text="주민등록번호")
            employee_id_label1.place(x=80, y=0)
            employee_rrn1_entry.append(Entry(subroot, width=7))
            employee_rrn1_entry[0].config(validate="key", validatecommand=(epmployee_entry_number_register(6),"%P"))
            employee_rrn1_entry[0].place(x=80, y=25)
            employee_id_label1 = Label(subroot, text="-")
            employee_id_label1.place(x=130, y=25)
            employee_rrn2_entry.append(Entry(subroot, width=8))
            employee_rrn2_entry[0].config(validate="key", validatecommand=(epmployee_entry_number_register(7),"%P"))
            employee_rrn2_entry[0].place(x=140, y=25)
            employee_YN_label.append(Label(subroot, text='지급여부'))
            employee_YN_label[0].place(x=200, y=0)
            employee_YN_radiobtn_var.append(IntVar())
            employee_YN_radiobtn1.append(Radiobutton(subroot, text="N", value=0, variable=employee_YN_radiobtn_var[0]))
            employee_YN_radiobtn1[0].place(x=200,y=25)
            employee_YN_radiobtn1[0].select()
            employee_YN_radiobtn2.append(Radiobutton(subroot, text="Y", value=1, variable=employee_YN_radiobtn_var[0]))
            employee_YN_radiobtn2[0].place(x=240,y=25)
            ##
            subroot_box_var.append(["","배우자","직계존속","직계비속"])
            subroot_box.append(ttk.Combobox(subroot, width=7, state='readonly', values=subroot_box_var[1]))
            subroot_box[1].current(0)
            subroot_box[1].place(x=0,y=75)
            employee_id_label2 = Label(subroot, text="주민등록번호")
            employee_id_label2.place(x=80, y=50)
            employee_rrn1_entry.append(Entry(subroot, width=7))
            employee_rrn1_entry[1].config(validate="key", validatecommand=(epmployee_entry_number_register(6),"%P"))
            employee_rrn1_entry[1].place(x=80, y=75)
            employee_id_label2 = Label(subroot, text="-")
            employee_id_label2.place(x=130, y=75)
            employee_rrn2_entry.append(Entry(subroot, width=8))
            employee_rrn2_entry[1].config(validate="key", validatecommand=(epmployee_entry_number_register(7),"%P"))
            employee_rrn2_entry[1].place(x=140, y=75)
            employee_YN_label.append(Label(subroot, text='지급여부'))
            employee_YN_label[1].place(x=200, y=50)
            employee_YN_radiobtn_var.append(IntVar())
            employee_YN_radiobtn1.append(Radiobutton(subroot, text="N", value=0, variable=employee_YN_radiobtn_var[1]))
            employee_YN_radiobtn1[1].place(x=200,y=75)
            employee_YN_radiobtn1[1].select()
            employee_YN_radiobtn2.append(Radiobutton(subroot, text="Y", value=1, variable=employee_YN_radiobtn_var[1]))
            employee_YN_radiobtn2[1].place(x=240,y=75)
            ##
            subroot_box_var.append(["","배우자","직계존속","직계비속"])
            subroot_box.append(ttk.Combobox(subroot, width=7, state='readonly', values=subroot_box_var[2]))
            subroot_box[2].current(0)
            subroot_box[2].place(x=0,y=125)
            employee_id_label3 = Label(subroot, text="주민등록번호")
            employee_id_label3.place(x=80, y=100)
            employee_rrn1_entry.append(Entry(subroot, width=7))
            employee_rrn1_entry[2].config(validate="key", validatecommand=(epmployee_entry_number_register(6),"%P"))
            employee_rrn1_entry[2].place(x=80, y=125)
            employee_id_label3 = Label(subroot, text="-")
            employee_id_label3.place(x=130, y=125)
            employee_rrn2_entry.append(Entry(subroot, width=8))
            employee_rrn2_entry[2].config(validate="key", validatecommand=(epmployee_entry_number_register(7),"%P"))
            employee_rrn2_entry[2].place(x=140, y=125)
            employee_YN_label.append(Label(subroot, text='지급여부'))
            employee_YN_label[2].place(x=200, y=100)
            employee_YN_radiobtn_var.append(IntVar())
            employee_YN_radiobtn1.append(Radiobutton(subroot, text="N", value=0, variable=employee_YN_radiobtn_var[2]))
            employee_YN_radiobtn1[2].place(x=200,y=125)
            employee_YN_radiobtn1[2].select()
            employee_YN_radiobtn2.append(Radiobutton(subroot, text="Y", value=1, variable=employee_YN_radiobtn_var[2]))
            employee_YN_radiobtn2[2].place(x=240,y=125)
            ##
            subroot_box_var.append(["","배우자","직계존속","직계비속"])
            subroot_box.append(ttk.Combobox(subroot, width=7, state='readonly', values=subroot_box_var[3]))
            subroot_box[3].current(0)
            subroot_box[3].place(x=0,y=175)
            employee_id_label4 = Label(subroot, text="주민등록번호")
            employee_id_label4.place(x=80, y=150)
            employee_rrn1_entry.append(Entry(subroot, width=7))
            employee_rrn1_entry[3].config(validate="key", validatecommand=(epmployee_entry_number_register(6),"%P"))
            employee_rrn1_entry[3].place(x=80, y=175)
            employee_id_label4 = Label(subroot, text="-")
            employee_id_label4.place(x=130, y=175)
            employee_rrn2_entry.append(Entry(subroot, width=8))
            employee_rrn2_entry[3].config(validate="key", validatecommand=(epmployee_entry_number_register(7),"%P"))
            employee_rrn2_entry[3].place(x=140, y=175)
            employee_YN_label.append(Label(subroot, text='지급여부'))
            employee_YN_label[3].place(x=200, y=150)
            employee_YN_radiobtn_var.append(IntVar())
            employee_YN_radiobtn1.append(Radiobutton(subroot, text="N", value=0, variable=employee_YN_radiobtn_var[3]))
            employee_YN_radiobtn1[3].place(x=200,y=175)
            employee_YN_radiobtn1[3].select()
            employee_YN_radiobtn2.append(Radiobutton(subroot, text="Y", value=1, variable=employee_YN_radiobtn_var[3]))
            employee_YN_radiobtn2[3].place(x=240,y=175)
            ##
            subroot_box_var.append(["","배우자","직계존속","직계비속"])
            subroot_box.append(ttk.Combobox(subroot, width=7, state='readonly', values=subroot_box_var[4]))
            subroot_box[4].current(0)
            subroot_box[4].place(x=0,y=225)
            employee_id_label5 = Label(subroot, text="주민등록번호")
            employee_id_label5.place(x=80, y=200)
            employee_rrn1_entry.append(Entry(subroot, width=7))
            employee_rrn1_entry[4].config(validate="key", validatecommand=(epmployee_entry_number_register(6),"%P"))
            employee_rrn1_entry[4].place(x=80, y=225)
            employee_id_label5 = Label(subroot, text="-")
            employee_id_label5.place(x=130, y=225)
            employee_rrn2_entry.append(Entry(subroot, width=8))
            employee_rrn2_entry[4].config(validate="key", validatecommand=(epmployee_entry_number_register(7),"%P"))
            employee_rrn2_entry[4].place(x=140, y=225)
            employee_YN_label.append(Label(subroot, text='지급여부'))
            employee_YN_label[4].place(x=200, y=200)
            employee_YN_radiobtn_var.append(IntVar())
            employee_YN_radiobtn1.append(Radiobutton(subroot, text="N", value=0, variable=employee_YN_radiobtn_var[4]))
            employee_YN_radiobtn1[4].place(x=200,y=225)
            employee_YN_radiobtn1[4].select()
            employee_YN_radiobtn2.append(Radiobutton(subroot, text="Y", value=1, variable=employee_YN_radiobtn_var[4]))
            employee_YN_radiobtn2[4].place(x=240,y=225)
            ##
            selection = employeelist_treeview.selection()
            to_view = employeelist_treeview.item(selection[0])['values']
            ind = [ind for ind, employee in enumerate(employeelist) if employee['id']==to_view[0]][0]
            try:
                for n, family in enumerate(employeelist[ind]["가족사항"]):
                    subroot_box[n].current(subroot_box_var[n].index(family["가족관계"]))
                    employee_rrn1_entry[n].delete(0,END)
                    employee_rrn1_entry[n].insert(0,family["주민등록번호"].split("-")[0])
                    employee_rrn2_entry[n].delete(0,END)
                    employee_rrn2_entry[n].insert(0,family["주민등록번호"].split("-")[1])
                    if family["지급여부"]:
                        employee_YN_radiobtn2[n].select()
                    else:
                        employee_YN_radiobtn1[n].select()
            except KeyError:
                pass
            subroot_btn = Button(subroot, text="저장하기", command=lambda:employee_family_apply())
            subroot_btn.place(x=280,y=270)
        def employee_family_apply():
            rrns = ["-".join([employee_rrn1_entry[n].get(),employee_rrn2_entry[n].get()]) for n in range(5)]
            if len([rrn for rrn in rrns if rrn!="-"]) != len(set([rrn for rrn in rrns if rrn!="-"])):
                showinfo("경고","주민등록번호가 중복되었습니다.")
                return None
            if [subroot_box[n].get() for n in range(5)].count("배우자") > 1:
                showinfo("경고","배우자가 둘 이상입니다.")
                return None
            if [subroot_box[n].get() for n in range(5)].count("직계존속") > 2:
                showinfo("경고","직계존속이 셋 이상입니다.")
                return None
            selection = employeelist_treeview.selection()
            to_mod = employeelist_treeview.item(selection[0])['values']
            ind = [ind for ind, employee in enumerate(employeelist) if employee['id']==to_mod[0]][0]
            employeelist[ind]["가족사항"] = [{"가족관계":subroot_box[n].get(), "주민등록번호":rrns[n], "지급여부":employee_YN_radiobtn_var[n].get()} for n in range(5) if rrns[n] != "-"]
            subroot.destroy()

    employee_family_btn = Button(root, text="가족수당", command=lambda: employee_family_add(root))
    employee_family_btn.place(x=670,y=50)


    employee_absence_label = Label(root, text="휴직")
    employee_absence_label.place(x=100, y=100+30)
    employee_absence_category_box_var = [""]+["육아휴직(첫째)","육아휴직(둘째)","육아휴직(셋째)","육아휴직(넷째)","질병휴직","유학휴직","기타휴직"]
    employee_absence_category_box = ttk.Combobox(root, height=5, width=12, state='readonly', values=employee_absence_category_box_var)
    employee_absence_category_box.current(0)
    employee_absence_category_box.place(x=140, y=100+30)
    employee_absence1_year_label = Label(root, text="년")
    employee_absence1_year_label.place(x=210+90+5, y=100+30)
    employee_absence1_year_box_var = [""]+[str(i) for i in range(datetime.now().year+2,2010,-1)]
    employee_absence1_year_box = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_absence1_year_box_var)
    employee_absence1_year_box.current(0)
    employee_absence1_year_box.place(x=160+90+5, y=100+30)
    employee_absence1_month_label = Label(root, text="월")
    employee_absence1_month_label.place(x=270+80+5, y=100+30)
    employee_absence1_month_box_var = [""]+[f"{i}" for i in range(1,13)]
    employee_absence1_month_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_absence1_month_box_var)
    employee_absence1_month_box.current(0)
    employee_absence1_month_box.place(x=230+85+5, y=100+30)
    def update_employee_absence1_day_box(event):
        if employee_absence1_year_box.get() == "":
            showinfo("경고", "연도를 선택해주십시오.")
            return None
        employee_absence1_day_box['values'] = [""] + [str(i) for i in range(1, (datetime(int(employee_absence1_year_box.get()), int(employee_absence1_month_box.get()), 1) + relativedelta(months=1) - relativedelta(days=1)).day+1)]
    employee_absence1_month_box.bind('<<ComboboxSelected>>', update_employee_absence1_day_box)
    employee_absence1_day_label = Label(root, text="일")
    employee_absence1_day_label.place(x=330+70+5, y=100+30)
    employee_absence1_day_box_var = [""]+[str(i) for i in range(1,32)]
    employee_absence1_day_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_absence1_day_box_var)
    employee_absence1_day_box.current(0)
    employee_absence1_day_box.place(x=290+75+5, y=100+30)
    employee_absence2_year_label = Label(root, text="~")
    employee_absence2_year_label.place(x=350-5+75, y=100+30)
    employee_absence2_year_label = Label(root, text="년")
    employee_absence2_year_label.place(x=420-5+70, y=100+30)
    employee_absence2_year_box_var = [""]+[str(i) for i in range(datetime.now().year+2,2010,-1)]
    employee_absence2_year_box = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_absence2_year_box_var)
    employee_absence2_year_box.current(0)
    employee_absence2_year_box.place(x=370-5+70, y=100+30)
    employee_absence2_month_label = Label(root, text="월")
    employee_absence2_month_label.place(x=480-5+60, y=100+30)
    employee_absence2_month_box_var = [""]+[f"{i}" for i in range(1,13)]
    employee_absence2_month_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_absence2_month_box_var)
    employee_absence2_month_box.current(0)
    employee_absence2_month_box.place(x=440-5+65, y=100+30)
    def update_employee_absence2_day_box(event):
        if employee_absence2_year_box.get() == "":
            showinfo("경고", "연도를 선택해주십시오.")
            return None
        employee_absence2_day_box['values'] = [""] + [str(i) for i in range(1, (datetime(int(employee_absence2_year_box.get()), int(employee_absence2_month_box.get()), 1) + relativedelta(months=1) - relativedelta(days=1)).day+1)]
    employee_absence2_month_box.bind('<<ComboboxSelected>>', update_employee_absence2_day_box)
    employee_absence2_day_label = Label(root, text="일")
    employee_absence2_day_label.place(x=540-5+50, y=100+30)
    employee_absence2_day_box_var = [""]+[str(i) for i in range(1,32)]
    employee_absence2_day_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_absence2_day_box_var)
    employee_absence2_day_box.current(0)
    employee_absence2_day_box.place(x=500-5+55, y=100+30)

    employee_contigous_radiobtn_var = IntVar()
    employee_contigous_radiobtn1 = Radiobutton(root, text="N", value=0, variable=employee_contigous_radiobtn_var)
    employee_contigous_radiobtn1.place(x=600,y=100+30)
    employee_contigous_radiobtn1.select()
    employee_contigous_radiobtn2 = Radiobutton(root, text="Y", value=1, variable=employee_contigous_radiobtn_var)
    employee_contigous_radiobtn2.place(x=635,y=100+30)

    def employee_absence_push(treeview):
        global employeelist_treeview
        global employeelist
        if employeelist_treeview.selection():
            category = employee_absence_category_box.get()
            try:
                datetime_start = datetime(int(employee_absence1_year_box.get()), int(employee_absence1_month_box.get()), int(employee_absence1_day_box.get()))
                datetime_end = datetime(int(employee_absence2_year_box.get()), int(employee_absence2_month_box.get()), int(employee_absence2_day_box.get()))
                contigous = employee_contigous_radiobtn_var.get()
            except ValueError:
                showinfo("경고", "날짜 형식이 맞지 않습니다.")
                return None
            if datetime_start >= datetime_end:
                showinfo("경고", "종료일은 시작일보다 빠를 수 없습니다.")
                return None
            if treeview.get_children():
                absence_list = [[category,datetime_start,datetime_end,contigous]]
                for item in treeview.get_children():
                    values = treeview.item(item)["values"]
                    values[1] = datetime.strptime(values[1], "%Y-%m-%d")
                    values[2] = datetime.strptime(values[2], "%Y-%m-%d")
                    absence_list.append(values)
                    treeview.delete(item)
                absence_list.sort(key=lambda x: x[2], reverse=True)
                [treeview.insert("",END,values=(category, datetime_start.date(), datetime_end.date(), contigous)) for category, datetime_start, datetime_end, contigous in absence_list]
            else:
                treeview.insert("",END,values=(category, datetime_start.date(), datetime_end.date(), contigous))

    employee_absence_pushbtn = Button(root, text="넣기", command=lambda: employee_absence_push(employee_absence_treeview))
    employee_absence_pushbtn.place(x=100,y=125+30)

    def employee_absence_pull(treeview):
        global employeelist
        selection = treeview.selection()
        if selection:
            to_pull = treeview.item(selection[0])['values']
            treeview.delete(selection)
    employee_absence_pullbtn = Button(root, text="뺴기", command=lambda: employee_absence_pull(employee_absence_treeview))
    employee_absence_pullbtn.place(x=100,y=155+30)


    def employee_absence_apply(treeview):
        to_apply = employeelist_treeview.item(employeelist_treeview.selection()[0])['values']
        ind = [ind for ind, employee in enumerate(employeelist) if employee['id']==to_apply[0]][0]
        absence_list = []
        for item in treeview.get_children():
            values = treeview.item(item)["values"]
            absence_list.append(values)
        absence_list.sort(key=lambda x: x[2], reverse=True)
        employeelist[ind]["휴직"] = absence_list
    employee_absence_applybtn = Button(root, text="적용", command=lambda: employee_absence_apply(employee_absence_treeview))
    employee_absence_applybtn.place(x=100,y=185+30)


    employee_absence_treeview = ttk.Treeview(root, height=8, columns=('휴직종류','시작일','종료일','아빠의달'), show="headings")
    employee_absence_treeview.heading("휴직종류",text="휴직종류")
    employee_absence_treeview.heading("시작일",text="시작일")
    employee_absence_treeview.heading("종료일",text="종료일")
    employee_absence_treeview.heading("아빠의달",text="아빠의달")
    # [employee_absence_treeview.insert("",END,values=(employee["id"],employee["성명"])) for employee in employeelist]
    employee_absence_treeview.column(0, width=120)
    employee_absence_treeview.column(1, width=160)
    employee_absence_treeview.column(2, width=160)
    employee_absence_treeview.column(3, width=55)
    employee_absence_treeview.place(x=140, y=125+30)


    def employee_career_entry_chk(input_name):
        if re.search("^[가-힣]*$", input_name):
            return True
        elif input_name =="":
            return True
        else:
            return False
    employee_career_entry_chkreg=root.register(employee_career_entry_chk)
    employee_career_label = Label(root, text="비교원\n경력")
    employee_career_label.place(x=0, y=75+40+300-10)
    employee_career_category_entry = Entry(root, width=12)
    employee_career_category_entry.config(validate="key", validatecommand=(employee_career_entry_chkreg,"%P"))
    employee_career_category_entry.place(x=40, y=75+40+300)
    employee_career_position_entry = Entry(root, width=12)
    employee_career_position_entry.config(validate="key", validatecommand=(employee_career_entry_chkreg,"%P"))
    employee_career_position_entry.place(x=140, y=75+40+300)
    employee_career_public_var = StringVar()
    employee_career_publicbtn1 = Radiobutton(root, text="N", value="비공무원", variable=employee_career_public_var)
    employee_career_publicbtn1.place(x=200+50,y=75+40+300)
    employee_career_publicbtn1.select()
    employee_career_publicbtn2 = Radiobutton(root, text="Y", value="공무원", variable=employee_career_public_var)
    employee_career_publicbtn2.place(x=240+50,y=75+40+300)
    def employee_career_rate_entry_chk(input_name):
        if re.search("^[0-9]{1,3}$", input_name) and int(input_name)<=100:
            return True
        elif input_name =="":
            return True
        else:
            return False
    employee_career_rate_entry_chkreg=root.register(employee_career_rate_entry_chk)
    employee_career_rate_entry = Entry(root, width=4)
    employee_career_rate_entry.config(validate="key", validatecommand=(employee_career_rate_entry_chkreg,"%P"))
    employee_career_rate_entry.place(x=350, y=75+40+300)
    employee_career_label = Label(root, text="%")
    employee_career_label.place(x=380, y=75+40+300)
    employee_career1_year_label = Label(root, text="년")
    employee_career1_year_label.place(x=210+80+90+80+5, y=75+40+300)
    employee_career1_year_box_var = [""]+[str(i) for i in range(datetime.now().year,1950,-1)]
    employee_career1_year_box = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_career1_year_box_var)
    employee_career1_year_box.current(0)
    employee_career1_year_box.place(x=160+80+90+80+5, y=75+40+300)
    employee_career1_month_label = Label(root, text="월")
    employee_career1_month_label.place(x=270+80+80+80+5, y=75+40+300)
    employee_career1_month_box_var = [""]+[f"{i}" for i in range(1,13)]
    employee_career1_month_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_career1_month_box_var)
    employee_career1_month_box.current(0)
    employee_career1_month_box.place(x=230+80+85+80+5, y=75+40+300)
    def update_employee_career1_day_box(event):
        if employee_career1_year_box.get() == "":
            showinfo("경고", "연도를 선택해주십시오.")
            return None
        employee_career1_day_box['values'] = [""] + [str(i) for i in range(1, (datetime(int(employee_career1_year_box.get()), int(employee_career1_month_box.get()), 1) + relativedelta(months=1) - relativedelta(days=1)).day+1)]
    employee_career1_month_box.bind('<<ComboboxSelected>>', update_employee_career1_day_box)
    employee_career1_day_label = Label(root, text="일")
    employee_career1_day_label.place(x=330+80+70+80+5, y=75+40+300)
    employee_career1_day_box_var = [""]+[str(i) for i in range(1,32)]
    employee_career1_day_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_career1_day_box_var)
    employee_career1_day_box.current(0)
    employee_career1_day_box.place(x=290+80+75+80+5, y=75+40+300)
    employee_career2_year_label = Label(root, text="~")
    employee_career2_year_label.place(x=350+80+75+80, y=75+40+300)
    employee_career2_year_label = Label(root, text="년")
    employee_career2_year_label.place(x=420+80+70+80, y=75+40+300)
    employee_career2_year_box_var = [""]+[str(i) for i in range(datetime.now().year,1950,-1)]
    employee_career2_year_box = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_career2_year_box_var)
    employee_career2_year_box.current(0)
    employee_career2_year_box.place(x=370+80+70+80, y=75+40+300)
    employee_career2_month_label = Label(root, text="월")
    employee_career2_month_label.place(x=480+80+60+80, y=75+40+300)
    employee_career2_month_box_var = [""]+[f"{i}" for i in range(1,13)]
    employee_career2_month_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_career2_month_box_var)
    employee_career2_month_box.current(0)
    employee_career2_month_box.place(x=440+80+65+80, y=75+40+300)
    def update_employee_career2_day_box(event):
        if employee_career2_year_box.get() == "":
            showinfo("경고", "연도를 선택해주십시오.")
            return None
        employee_career2_day_box['values'] = [""] + [str(i) for i in range(1, (datetime(int(employee_career2_year_box.get()), int(employee_career2_month_box.get()), 1) + relativedelta(months=1) - relativedelta(days=1)).day+1)]
    employee_career2_month_box.bind('<<ComboboxSelected>>', update_employee_career2_day_box)
    employee_career2_day_label = Label(root, text="일")
    employee_career2_day_label.place(x=540+80+50+80, y=75+40+300)
    employee_career2_day_box_var = [""]+[str(i) for i in range(1,32)]
    employee_career2_day_box = ttk.Combobox(root, height=5, width=2, state='readonly', values=employee_career2_day_box_var)
    employee_career2_day_box.current(0)
    employee_career2_day_box.place(x=500+80+55+80, y=75+40+300)


    def employee_career_push(treeview):
        global employeelist_treeview
        global employeelist
        if employeelist_treeview.selection():
            category = employee_career_category_entry.get()
            position = employee_career_position_entry.get()
            public = employee_career_public_var.get()
            rate = employee_career_rate_entry.get()
            try:
                datetime_start = datetime(int(employee_career1_year_box.get()), int(employee_career1_month_box.get()), int(employee_career1_day_box.get()))
                datetime_end = datetime(int(employee_career2_year_box.get()), int(employee_career2_month_box.get()), int(employee_career2_day_box.get()))
            except ValueError:
                showinfo("경고", "날짜 형식이 맞지 않습니다.")
                return None
            if datetime_start >= datetime_end:
                showinfo("경고", "종료일은 시작일보다 같거나 빠를 수 없습니다.")
                return None
            if treeview.get_children():
                career_list = [[category, position, public, rate, datetime_start.strftime("%Y-%m-%d"), datetime_end.strftime("%Y-%m-%d")]]
                for item in treeview.get_children():
                    values = treeview.item(item)["values"]
                    values[4] = values[4]
                    values[5] = values[5]
                    career_list.append(values)
                    treeview.delete(item)
                career_list.sort(key=lambda x: x[5], reverse=True)
                [treeview.insert("",END,values=(category, position, public, rate, datetime_start, datetime_end)) for category, position, public, rate, datetime_start, datetime_end in career_list]
            else:
                treeview.insert("",END,values=(category, position, public, rate, datetime_start.strftime("%Y-%m-%d"), datetime_end.strftime("%Y-%m-%d")))
            employee_career_category_entry.delete(0,END)
            employee_career_position_entry.delete(0,END)
            employee_career_publicbtn1.select()
            employee_career_rate_entry.delete(0,END)
            employee_career1_year_box.current(0)
            employee_career1_month_box.current(0)
            employee_career1_day_box.current(0)
            employee_career2_year_box.current(0)
            employee_career2_month_box.current(0)
            employee_career2_day_box.current(0)
    employee_career_pushbtn = Button(root, text="넣기", command=lambda: employee_career_push(employee_career_treeview))
    employee_career_pushbtn.place(x=0,y=100+40+300)

    def employee_career_pull(treeview):
        global employeelist
        selection = treeview.selection()
        if selection:
            to_pull = treeview.item(selection[0])['values']
            treeview.delete(selection)
    employee_career_pullbtn = Button(root, text="뺴기", command=lambda: employee_career_pull(employee_career_treeview))
    employee_career_pullbtn.place(x=0,y=130+40+300)


    def employee_career_apply(treeview):
        to_apply = employeelist_treeview.item(employeelist_treeview.selection()[0])['values']
        ind = [ind for ind, employee in enumerate(employeelist) if employee['id']==to_apply[0]][0]
        career_list = []
        for item in treeview.get_children():
            values = treeview.item(item)["values"]
            career_list.append(values)
        career_list.sort(key=lambda x: x[5], reverse=True)
        employeelist[ind]["경력"] = career_list
    employee_career_applybtn = Button(root, text="적용", command=lambda: employee_career_apply(employee_career_treeview))
    employee_career_applybtn.place(x=0,y=160+40+300)

    employee_career_treeview = ttk.Treeview(root, height=5, columns=('근무처','직위','공무원여부','환산율','시작일','종료일'), show="headings")
    employee_career_treeview.heading("근무처",text="근무처")
    employee_career_treeview.heading("직위",text="직위")
    employee_career_treeview.heading("공무원여부",text="공무원여부")
    employee_career_treeview.heading("환산율",text="환산율")
    employee_career_treeview.heading("시작일",text="시작일")
    employee_career_treeview.heading("종료일",text="종료일")
    employee_career_treeview.column(0, width=100)
    employee_career_treeview.column(1, width=100)
    employee_career_treeview.column(2, width=100)
    employee_career_treeview.column(3, width=60)
    employee_career_treeview.column(4, width=180)
    employee_career_treeview.column(5, width=180)
    employee_career_treeview.place(x=40, y=100+40+300)

    employee_elder_candidate_text = Text(root, width = 103, height = 4)
    employee_elder_candidate_text.place(x=40, y=570)


    def employeelist_add(treeview):
        global employeelist
        def employeelist_information_insert(ind):
            employeelist.insert(ind,{"id":employee_id,
                                     "성명":employee_name_entry.get(),
                                     "주민번호":"-".join([employee_rrn1_entry.get(), employee_rrn2_entry.get()]),
                                     "직종":employee_category_radiobtn_var.get(),
                                     "승급년월일":"-".join([employee_year_pay_increasebtn.get().replace("년",""), employee_month_pay_increasebtn.get().replace("월",""), employee_day_pay_increasebtn.get().replace("일","")]),
                                     "근무연한":[int(employee_years_of_service_year_entry.get()),int(employee_years_of_service_month_entry.get()),int(employee_years_of_service_day_entry.get())],
                                     "호봉":int(employee_step_entry.get()),
                                     "급":employee_gradebtn.get(),
                                     "현근무년수변경일":"" if employee_year_datechangebtn.get() == "" or employee_month_datechangebtn.get() == "" or employee_day_datechangebtn.get() == "" else "-".join([employee_year_datechangebtn.get().replace("년",""), employee_month_datechangebtn.get().replace("월",""), employee_day_datechangebtn.get().replace("일","")]),
                                     "현부서임용일":"" if employee_appointment_year_box.get() == "" or employee_appointment_month_box.get() == "" or employee_appointment_day_box.get() == "" else "-".join([employee_appointment_year_box.get(), employee_appointment_month_box.get(), employee_appointment_day_box.get()]),
                                     "계속근무여부":employee_keep_working_radiobtn_var.get(),
                                     "퇴직일":"" if employee_retire_year_box.get() == "" or employee_retire_month_box.get() == "" or employee_retire_day_box.get() == "" else "-".join([employee_retire_year_box.get(), employee_retire_month_box.get(), employee_retire_day_box.get()]),
                                     "보직":employee_position_btn.get(),
                                     "가산정원":employee_special_class_btn.get(),
                                     "원로교사":employee_elder_radiobtn_var.get(),
                                     "연가보상일수":employee_anual_leave_compensation_entry.get()
                                     })
        if employee_id_entry.get() and employee_name_entry.get() and employee_rrn1_entry.get() and employee_rrn2_entry.get():
            employee_id = int(employee_id_entry.get())
            if employee_id in [employee["id"] for employee in employeelist]:
                employee_id = max([employee["id"] for employee in employeelist])+1
                employee_id_entry.delete(0,END)
                employee_id_entry.insert(0,employee_id)
            if len(treeview.selection()) == 0:
                treeview.insert("", END, values=(employee_id, employee_name_entry.get()))
                employeelist_information_insert(len(employeelist))
            else:
                selection = treeview.selection()
                ind = treeview.get_children().index(selection[0])
                if ind == len(treeview.get_children())-1:
                    treeview.insert("", END, values=(employee_id, employee_name_entry.get()))
                else:
                    treeview.insert("", ind+1, values=(employee_id, employee_name_entry.get()))
                employeelist_information_insert(ind+1)
            employee_id_entry.delete(0,END)
            employee_name_entry.delete(0,END)
            employee_rrn1_entry.delete(0,END)
            employee_rrn2_entry.delete(0,END)
            employee_years_of_service_year_entry.delete(0,END)
            employee_step_entry.delete(0,END)
            employee_gradebtn.current(0)
            if employee_retire_year_box.get() == "" or employee_retire_month_box.get() == "" or employee_retire_day_box.get() == "":
                employee_retire_year_box.current(0)
                employee_retire_month_box.current(0)
                employee_retire_day_box.current(0)
            treeview.selection_remove(selection)

    employeelist_addbtn = Button(root, text="추가", command=lambda: employeelist_add(employeelist_treeview))
    employeelist_addbtn.place(x=760,y=30+70)

    def employeelist_del(treeview):
        global employeelist
        selection = treeview.selection()
        if selection:
            to_del = treeview.item(selection[0])['values']
            treeview.delete(selection[0])
            employeelist = [employee for employee in employeelist if employee['id']!=to_del[0]]

    employeelist_delbtn = Button(root, text="삭제", command=lambda: employeelist_del(employeelist_treeview))
    employeelist_delbtn.place(x=760,y=60+70)

    def employeelist_mod(treeview):
        global employeelist
        selection = treeview.selection()
        if selection:
            to_mod = treeview.item(selection[0])['values']
            ind = [ind for ind, employee in enumerate(employeelist) if employee['id']==to_mod[0]][0]
            employeelist[ind]["id"]=to_mod[0]
            employeelist[ind]["성명"]=employee_name_entry.get()
            employeelist[ind]["주민번호"]="-".join([employee_rrn1_entry.get(), employee_rrn2_entry.get()])
            employeelist[ind]["직종"]=employee_category_radiobtn_var.get()
            employeelist[ind]["승급년월일"]="-".join([employee_year_pay_increasebtn.get().replace("년",""), employee_month_pay_increasebtn.get().replace("월",""), employee_day_pay_increasebtn.get().replace("일","")])
            employeelist[ind]["근무연한"]=[int(employee_years_of_service_year_entry.get()),int(employee_years_of_service_month_entry.get()),int(employee_years_of_service_day_entry.get())]
            employeelist[ind]["호봉"]=int(employee_step_entry.get())
            employeelist[ind]["급"]=employee_gradebtn.get()
            employeelist[ind]["현근무년수변경일"]="" if employee_year_datechangebtn.get() == "" or employee_month_datechangebtn.get() == "" or employee_day_datechangebtn.get() == "" else "-".join([employee_year_datechangebtn.get().replace("년",""), employee_month_datechangebtn.get().replace("월",""), employee_day_datechangebtn.get().replace("일","")])
            employeelist[ind]["현부서임용일"]="" if employee_appointment_year_box.get() == "" or employee_appointment_month_box.get() == "" or employee_appointment_day_box.get() == "" else "-".join([employee_appointment_year_box.get(), employee_appointment_month_box.get(), employee_appointment_day_box.get()])
            employeelist[ind]["계속근무여부"]=employee_keep_working_radiobtn_var.get()
            employeelist[ind]["퇴직일"]="" if employee_retire_year_box.get() == "" or employee_retire_month_box.get() == "" or employee_retire_day_box.get() == "" else "-".join([employee_retire_year_box.get(), employee_retire_month_box.get(), employee_retire_day_box.get()])
            employeelist[ind]["보직"]=employee_position_btn.get()
            employeelist[ind]["가산정원"]=employee_special_class_btn.get()
            employeelist[ind]["원로교사"]=employee_elder_radiobtn_var.get()
            employeelist[ind]["연가보상일수"]=employee_anual_leave_compensation_entry.get()
            employeelist[ind]["감봉시작일"]="" if employee_paycut_start_year_box.get() == "" or employee_paycut_start_month_box.get() == "" or employee_paycut_start_day_box.get() == "" else "-".join([employee_paycut_start_year_box.get(), employee_paycut_start_month_box.get(), employee_paycut_start_day_box.get()])
            employeelist[ind]["감봉종료일"]="" if employee_paycut_end_year_box.get() == "" or employee_paycut_end_month_box.get() == "" or employee_paycut_end_day_box.get() == "" else "-".join([employee_paycut_end_year_box.get(), employee_paycut_end_month_box.get(), employee_paycut_end_day_box.get()])
            employeelist[ind]["감봉율"]=employee_paycut_rate_entry.get()
            employeelist[ind]["승급제한시작일"]="" if employee_upgrade_restriction_start_year_box.get() == "" or employee_upgrade_restriction_start_month_box.get() == "" or employee_upgrade_restriction_start_day_box.get() == "" else "-".join([employee_upgrade_restriction_start_year_box.get(), employee_upgrade_restriction_start_month_box.get(), employee_upgrade_restriction_start_day_box.get()])
            employeelist[ind]["승급제한종료일"]="" if employee_upgrade_restriction_end_year_box.get() == "" or employee_upgrade_restriction_end_month_box.get() == "" or employee_upgrade_restriction_end_day_box.get() == "" else "-".join([employee_upgrade_restriction_end_year_box.get(), employee_upgrade_restriction_end_month_box.get(), employee_upgrade_restriction_end_day_box.get()])
            treeview.delete(selection)
            treeview.insert("", ind, values=(employeelist[ind]["id"],employeelist[ind]["성명"]))
    employeelist_modbtn = Button(root, text="수정", command=lambda: employeelist_mod(employeelist_treeview))
    employeelist_modbtn.place(x=760,y=90+70)

    employeelist_chkbtn = Button(root, text="확인", command=lambda: print(employeelist))
    employeelist_chkbtn.place(x=760,y=120+70)


    ############### Salary ###############
    config = {"근속가봉표" : {"2022":72900,"2021":71700,"2020":71000,"2019":68800,"2018":67400,"2017":65500,"2016":63200},
              "가산금" : [ele for ele in [0, 50000, 60000, 80000, 100000, 100000] for i in range(5)],
              "추가가산금" : [ele for ele in [0, 0, 0, 0, 10000, 30000] for i in range(5)],
              "직급보조비" : {"2022":{"5급":250000,"6급":175000,"7급":165000,"8급":155000,"9급":155000,"교장":400000,"교감":250000}, "2021":{"5급":250000,"6급":165000,"7급":155000,"8급":145000,"9급":145000,"교장":400000,"교감":250000}},
              "급식비" : {"2022":140000,"2021":140000,"2020":140000,"2019":130000,"2018":130000,"2017":130000,"2016":130000},
              "보전수당" : {"2022":{"교장":70000,"교감":10000},"2021":{"교장":70000,"교감":10000}},
              "교직수당" : {"2022":250000,"2021":250000},
              "원로교사수당" : {"2022":50000,"2021":50000},
              "부장교사수당" : {"2022":70000,"2021":70000},
              "담임교사수당" : {"2022":130000,"2021":130000},
              "보건교사수당" : {"2022":30000,"2021":30000},
              "상담교사수당" : {"2022":20000,"2021":20000},
              "가족수당" : {"2022":{"배우자":40000, "직계존속":20000, "직계비속":[20000, 60000, 100000, 100000, 100000]}, "2021":{"배우자":40000, "직계존속":20000, "직계비속":[20000, 60000, 100000, 100000, 100000]}},
              "시간외근무수당" : {"2022":{"행정직":{"9급":9032,"8급":9992,"7급":11130,"6급":12321,"5급":14446},"교장":0,"교감":14434,"교원":[(1,19,11514),(20,29,12790),(30,50,13730)]},
                  "2021":{"행정직":{"9급":8887,"8급":9832,"7급":10952,"6급":12124,"5급":14215},"교장":0,"교감":14434,"교원":[(1,19,11330),(20,29,12585),(30,50,13511)]}},
              "학교운영수당" : {"2022":30000,"2021":30000,"2020":30000,"2019":30000,"2018":30000,"2017":30000,"2016":30000},
              "육아휴직수당" : {"2022":((0.8, 1500000, 700000),(0.8, 1500000, 700000)),"2021":((0.8, 1500000, 700000),(0.8, 1200000, 700000))},
              "명절년월일" : {"2022":{"추석":"2022-9-10", "설":"2023-1-22"}, "2021":{"추석":"2021-9-21", "설":"2022-2-1"}}}

    home = os.path.expanduser("~")
    config_dir = os.path.join(home, ".EmploymentCostPreCalculationProgramForSchool")
    config_path = os.path.join(config_dir, "config.json")
    if os.path.exists(config_path):
        with open(config_path, "r", encoding='UTF-8') as f:
            config = json.load(f)
    else:
        if not os.path.isdir(config_dir):
            os.mkdir(config_dir)
        with open(config_path, "w", encoding='UTF-8') as f:
            json.dump(config, f, indent=4, ensure_ascii=False)

    근속가봉표 = config["근속가봉표"]
    가산금 = config["가산금"]
    추가가산금 = config["추가가산금"]
    직급보조비 = config["직급보조비"]
    급식비 = config["급식비"]
    보전수당 = config["보전수당"]
    교직수당 = config["교직수당"]
    원로교사수당 = config["원로교사수당"]
    부장교사수당 = config["부장교사수당"]
    담임교사수당 = config["담임교사수당"]
    보건교사수당 = config["보건교사수당"]
    상담교사수당 = config["상담교사수당"]
    가족수당 = config["가족수당"]
    시간외근무수당 = config["시간외근무수당"]
    학교운영수당 = config["학교운영수당"]
    육아휴직수당 = config["육아휴직수당"]
    명절년월일 = config["명절년월일"]

    def 본봉표(연도, 근속가봉표={}):
        url = 'http://www.mpm.go.kr/mpm/info/resultPay/bizSalary/'+연도
        html = urlopen(url)

        bsObject = BeautifulSoup(html, "html.parser")

        # remove <div class="mgt10" id="pay2020_1_1" style="display:none">
        trashtags = bsObject.findAll('div', attrs={'class':'mgt10'})
        [trashtag.extract() for trashtag in trashtags]

        tag = bsObject.findAll('div', attrs={'class':"table-responsive"})
        tables = pd.read_html(str(tag))

        # 표 정리
        tables[0].columns=['호봉','1급','2급','3급','4급','5급','6급','7급','8급','9급']
        tables[3].columns=['호봉','봉급','호봉','봉급']
        tables[3]=pd.concat([tables[3].iloc[:,0:2],tables[3].iloc[:,2:4]])
        tables[3]=pd.concat([tables[3], pd.DataFrame(zip(range(41,51), [tables[3].iloc[-1,1]+x*(n+1) for n, x in enumerate([근속가봉표[연도]]*10)]),columns=['호봉','봉급'])],ignore_index=True, sort=False)

        return tables

    def rrn_to_datetime(rrn):
        if int(rrn.split("-")[1][0]) in [3,4]:
            return datetime(int("20"+rrn.split()[0][:2]),int(rrn.split("-")[0][2:4]),int(rrn.split("-")[0][4:6]))
        else:
            return datetime(int("19"+rrn.split()[0][:2]),int(rrn.split("-")[0][2:4]),int(rrn.split("-")[0][4:6]))

    def merge_date_ranges(data):
        result = []
        t_old = data[0]
        for t in data[1:]:
            if t_old[1] >= t[0]:  #I assume that the data is sorted already
                t_old = ((min(t_old[0], t[0]), max(t_old[1], t[1])))
            else:
                result.append(t_old)
                t_old = t
        else:
            result.append(t_old)
        return result

    class 급여생성:
        본봉표 = None
        def __init__(self, 교직원, 작업연도, 명절년월일):
            self.교직원 = 교직원
            self.명절월 = list(map(lambda x : int(re.sub("^2$", "14", re.sub("^1$", "13", x.split("-")[1]))), 명절년월일.values()))
            self.급여 = pd.DataFrame()
            self.작업연도 = 작업연도
        def strptime(self, str_date):
            return datetime.strptime(str_date, "%Y-%m-%d")

        def 호봉(self, working_month):
            if self.교직원["직종"] == "기간제교원":
                전기승급일 = self.strptime(self.교직원['승급년월일']) + relativedelta(years=-1, months=self.교직원['근무연한'][1], days=self.교직원['근무연한'][2])
                호봉 = self.교직원["호봉"]+(datetime(int(작업연도), 3, 1) - 전기승급일).days//365.25
            else:
                if '휴직' in self.교직원.keys():
                    all_sick_leaves = [(self.strptime(date_start), self.strptime(date_end)) for category, date_start, date_end, _ in self.교직원["휴직"] if re.search("질병휴직", category)]
                else:
                    all_sick_leaves = []
                if '승급제한시작일' in self.교직원.keys():
                    upgrade_restriction_duration = (self.strptime(self.교직원['승급제한시작일']), self.strptime(self.교직원['승급제한종료일']))
                else:
                    upgrade_restriction_duration = []
                all_upgrade_restriction_durations = all_sick_leaves + [upgrade_restriction_duration]
                all_upgrade_restriction_durations = merge_date_ranges(all_upgrade_restriction_durations)
                if all_upgrade_restriction_durations != [[]] and [num for num, dates in enumerate(all_upgrade_restriction_durations) if dates[0] <= working_month and working_month < dates[1]] != []:
                    num = [num for num, dates in enumerate(all_upgrade_restriction_durations) if dates[0] <= working_month and working_month < dates[1]][0]
                    full_upgrade_restriction_duration = working_month - all_upgrade_restriction_durations[num][0]
                    for i in range(num):
                        full_upgrade_restriction_duration += all_upgrade_restriction_durations[i][1] - all_upgrade_restriction_durations[i][0]
                    호봉 = self.교직원['호봉']+(working_month + relativedelta(months=self.교직원['근무연한'][1], days=self.교직원['근무연한'][2]) - self.strptime(self.교직원['승급년월일']) - full_upgrade_restriction_duration).days//365.25
                elif all_upgrade_restriction_durations != [[]] and [num for num, dates in enumerate(all_upgrade_restriction_durations) if dates[1] <= working_month] != []:
                    num = max([num for num, dates in enumerate(all_upgrade_restriction_durations) if dates[1] <= working_month])
                    full_upgrade_restriction_duration = all_upgrade_restriction_durations[num][1] - all_upgrade_restriction_durations[num][0]
                    for i in range(num):
                        full_upgrade_restriction_duration += all_upgrade_restriction_durations[i][1] - all_upgrade_restriction_durations[i][0]
                    호봉 = self.교직원['호봉']+(working_month + relativedelta(months=self.교직원['근무연한'][1], days=self.교직원['근무연한'][2]) - self.strptime(self.교직원['승급년월일']) - full_upgrade_restriction_duration).days//365.25
                else:
                    호봉 = self.교직원['호봉']+(working_month + relativedelta(months=self.교직원['근무연한'][1], days=self.교직원['근무연한'][2]) - self.strptime(self.교직원['승급년월일'])).days//365.25
            return 호봉

        def 본봉(self):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and working_month < self.strptime(self.교직원['현부서임용일']):
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            호봉 = self.호봉(working_month)
            print(호봉)
            if self.교직원['급'] == '':
                if 호봉 > 50:
                    return int(self.본봉표[3][self.본봉표[3]['호봉']==50]['봉급'])
                return int(self.본봉표[3][self.본봉표[3]['호봉']==호봉]['봉급'])
            else:
                return int(self.본봉표[0][self.본봉표[0]['호봉']==호봉][self.교직원['급']])

        def 정근수당(self):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            비공무원경력일수 = 0
            if "경력" in self.교직원.keys():
                for 비공무원경력 in filter(lambda x:x[2]=="비공무원", self.교직원["경력"]):
                    비공무원경력일수 += (((self.strptime(비공무원경력[5])-self.strptime(비공무원경력[4])).days+1))*비공무원경력[3]/100
            근무연한일수 = (self.strptime(self.교직원["현근무년수변경일"]) + relativedelta(years=self.교직원['근무연한'][0],months=self.교직원['근무연한'][1],days=self.교직원['근무연한'][2]) - self.strptime(self.교직원["현근무년수변경일"])).days
            추가근무연한일수 = (working_month - self.strptime(self.교직원["현근무년수변경일"])).days + 1
            # 근무연수 = int((근무연한일수 - 비공무원경력일수 + 추가근무연한일수)//365)
            근무연수 = int((근무연한일수 + 추가근무연한일수)//365)
            real_working_months = 6
            if self.교직원['현부서임용일'] != "":
                if working_month - relativedelta(months=6) <= self.strptime(self.교직원["현부서임용일"]):
                    if self.strptime(self.교직원["현부서임용일"]).day == 1:
                        real_working_months = (working_month - self.strptime(self.교직원["현부서임용일"])).days//30
                    else:
                        real_working_months = (working_month - self.strptime(self.교직원["현부서임용일"])).days//30 - 1
            if self.현재월 == 7:
                working_total_days_list = []
                for i in range(0,6):
                    working_month_temp = (datetime(int(작업연도), 1, 1)+relativedelta(months=i))                    
                    working_month_totaldays = (working_month_temp+relativedelta(months=1)-relativedelta(days=1)).day
                    prolation = 0
                    if "휴직" in self.교직원.keys():                        
                        ongoing_leaves = [[category, self.strptime(date_start), self.strptime(date_end)] for category, date_start, date_end, _ in self.교직원["휴직"] if ((self.strptime(date_start)<=working_month_temp+relativedelta(months=1)-relativedelta(days=1) and working_month_temp<=self.strptime(date_start)) or (self.strptime(date_end)<=working_month_temp+relativedelta(months=1)-relativedelta(days=1) and working_month_temp<=self.strptime(date_end)) or (working_month_temp<=self.strptime(date_end) and working_month_temp>=self.strptime(date_start)))]
                        for ongoing_leave in ongoing_leaves:                            
                            prolation1, prolation2 = (working_month_temp+relativedelta(months=1)-ongoing_leave[1]).days, (ongoing_leave[2]-working_month_temp).days+1
                            if prolation1 <= working_month_totaldays and prolation2 <= working_month_totaldays:
                                prolation = prolation1 + prolation2 - working_month_totaldays
                            elif prolation1 <= working_month_totaldays:
                                prolation = prolation1
                            elif prolation2 <= working_month_totaldays:
                                prolation = prolation2
                            else:
                                prolation = working_month_totaldays
                            days = (working_month_temp - ongoing_leave[1]).days + 1
                            prolation =+ prolation
                            # 현재 휴직중인 경우 정근수당 미지급
                            if (working_month - ongoing_leave[1]).days >= 0 and (ongoing_leave[2] - working_month).days > 0 :
                                return int(0)
                # 지급대상기간 중 정직 또는 강등인 경우 정근수당 미지급
                if "감봉시작일" in self.교직원.keys() and self.교직원["감봉율"] == "100":
                    if datetime(int(작업연도), 1, 1) <= self.strptime(self.교직원["감봉시작일"]) and self.strptime(self.교직원["감봉시작일"]) <= datetime(int(작업연도), 1, 1) + relativedelta(months=6)-relativedelta(days=1):
                        return int(0)
                    working_total_days_list.append(working_month_totaldays - prolation)
                real_working_months -= sum([1 for working_total_days in working_total_days_list if working_total_days < 30])
            elif self.현재월 == 13:
                working_total_days_list = []
                for i in range(6,12):
                    working_month_temp = (datetime(int(작업연도), 1, 1)+relativedelta(months=i))                    
                    working_month_totaldays = (working_month_temp+relativedelta(months=1)-relativedelta(days=1)).day
                    prolation = 0
                    if "휴직" in self.교직원.keys():
                        ongoing_leaves = [[category, self.strptime(date_start), self.strptime(date_end)] for category, date_start, date_end, _ in self.교직원["휴직"] if ((self.strptime(date_start)<=working_month_temp+relativedelta(months=1)-relativedelta(days=1) and working_month_temp<=self.strptime(date_start)) or (self.strptime(date_end)<=working_month_temp+relativedelta(months=1)-relativedelta(days=1) and working_month_temp<=self.strptime(date_end)) or (working_month_temp<=self.strptime(date_end) and working_month_temp>=self.strptime(date_start)))]
                        for ongoing_leave in ongoing_leaves:                            
                            prolation1, prolation2 = (working_month_temp+relativedelta(months=1)-ongoing_leave[1]).days, (ongoing_leave[2]-working_month_temp).days+1
                            if prolation1 <= working_month_totaldays and prolation2 <= working_month_totaldays:
                                prolation = prolation1 + prolation2 - working_month_totaldays
                            elif prolation1 <= working_month_totaldays:
                                prolation = prolation1
                            elif prolation2 <= working_month_totaldays:
                                prolation = prolation2
                            else:
                                prolation = working_month_totaldays
                            days = (working_month_temp - ongoing_leave[1]).days + 1
                            prolation =+ prolation
                            # 현재 휴직중인 경우 정근수당 미지급
                            if (working_month - ongoing_leave[1]).days >= 0 and (ongoing_leave[2] - working_month).days > 0 :
                                return int(0)
                # 지급대상기간 중 정직 또는 강등인 경우 정근수당 미지급
                if "감봉시작일" in self.교직원.keys() and self.교직원["감봉율"] == "100":
                    if datetime(int(작업연도), 1, 1) + relativedelta(months=6) <= self.strptime(self.교직원["감봉시작일"]) and self.strptime(self.교직원["감봉시작일"]) <= datetime(int(작업연도), 1, 1) + relativedelta(months=12)-relativedelta(days=1):
                        return int(0)
                    working_total_days_list.append(working_month_totaldays - prolation)
                real_working_months -= sum([1 for working_total_days in working_total_days_list if working_total_days < 30])
            if self.현재월 in [13,7]:
                if self.교직원["직종"] == "기간제교원":
                    if self.교직원["근무연한"][0] + 추가근무연한일수//365 >= 10:
                        if "계속근무여부" in self.교직원.keys() and self.교직원["계속근무여부"] == 1:
                            return int(self.본봉()*0.05*10//10*10)
                        return int(self.본봉()*0.05*10*real_working_months/6//10*10)
                    else:
                        if "계속근무여부" in self.교직원.keys() and self.교직원["계속근무여부"] == 1:
                            int(self.본봉()*0.05*(self.교직원["근무연한"][0] + 추가근무연한일수//365)//10*10)
                        return int(self.본봉()*0.05*(self.교직원["근무연한"][0] + 추가근무연한일수//365)*real_working_months/6//10*10)
                else:
                    if 근무연수 >= 10:
                        return int(self.본봉()*0.05*10*real_working_months/6//10*10)
                    elif 근무연수 <= 0:
                        return int(0)
                    else:
                        return int(self.본봉()*0.05*근무연수*real_working_months/6//10*10)
            else:
                return int(0)

        def 정근수당가산금(self, 가산금):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            비공무원경력일수 = 0
            if "경력" in self.교직원.keys():
                for 비공무원경력 in filter(lambda x:x[2]=="비공무원", self.교직원["경력"]):
                    비공무원경력일수 += ((self.strptime(비공무원경력[5])-self.strptime(비공무원경력[4])).days+1)*비공무원경력[3]/100
            근무연한일수 = (self.strptime(self.교직원["현근무년수변경일"])+relativedelta(years=self.교직원['근무연한'][0],months=self.교직원['근무연한'][1],days=self.교직원['근무연한'][2])-self.strptime(self.교직원["현근무년수변경일"])).days
            추가근무연한일수 = (working_month-self.strptime(self.교직원["현근무년수변경일"])).days+1
            # 근무연수 = int((근무연한일수 - 비공무원경력일수 + 추가근무연한일수)//365)
            근무연수 = int((근무연한일수 + 추가근무연한일수)//365)
            if self.교직원["직종"] == "기간제교원":
                if (self.교직원["근무연한"][0] + 추가근무연한일수//365)>25:
                    return 가산금[-1]
                else:
                    return 가산금[(self.교직원["근무연한"][0] + 추가근무연한일수//365)]
            else:
                if 근무연수>25:
                    return 가산금[-1]
                elif 근무연수<0:
                    return 가산금[0]
                else:
                    return 가산금[근무연수]

        def 정근수당추가가산금(self, 추가가산금):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            비공무원경력일수 = 0
            if "경력" in self.교직원.keys():
                for 비공무원경력 in filter(lambda x:x[2]=="비공무원", self.교직원["경력"]):
                    비공무원경력일수 += ((self.strptime(비공무원경력[5])-self.strptime(비공무원경력[4])).days+1)*비공무원경력[3]/100
            근무연한일수 = (self.strptime(self.교직원["현근무년수변경일"])+relativedelta(years=self.교직원['근무연한'][0],months=self.교직원['근무연한'][1],days=self.교직원['근무연한'][2])-self.strptime(self.교직원["현근무년수변경일"])).days
            추가근무연한일수 = (working_month-self.strptime(self.교직원["현근무년수변경일"])).days+1
            # 근무연수 = int((근무연한일수 - 비공무원경력일수 + 추가근무연한일수)//365)
            근무연수 = int((근무연한일수 + 추가근무연한일수)//365)
            if self.교직원["직종"] == "기간제교원":
                if (self.교직원["근무연한"][0] + 추가근무연한일수//365)>25:
                    return 추가가산금[-1]
                else:
                    return 추가가산금[(self.교직원["근무연한"][0] + 추가근무연한일수//365)]
            else:
                if 근무연수>25:
                    return 추가가산금[-1]
                elif 근무연수<0:
                    return 추가가산금[0]
                else:
                    return 추가가산금[근무연수]

        def 정액급식비(self, 급식비):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            return 급식비

        def 직급보조비(self, 직급보조비):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            if self.교직원['직종'] == '행정직':
                return 직급보조비[self.작업연도][self.교직원['급']]
            elif self.교직원['보직'] == '교장':
                return 직급보조비[self.작업연도][self.교직원['보직']]
            elif self.교직원['보직'] == '교감':
                return 직급보조비[self.작업연도][self.교직원['보직']]
            else:
                return int(0)

        def 명절휴가비(self):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            if self.현재월 in self.명절월:
                return int(self.본봉()*0.6//10*10)
            else:
                return int(0)

        def 관리업무수당(self):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            if self.교직원["보직"] in ["교장"]:
                return int(self.본봉()*0.078//10*10)
            else:
                return int(0)

        def 보전수당(self, 보전수당):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            if self.교직원["보직"] in ["교장", "교감"]:
                return 보전수당[self.작업연도][self.교직원["보직"]]
            else:
                return int(0)

        def 교직수당(self, 교직수당):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            if self.교직원["직종"] in ["교원","기간제교원"]:
                return 교직수당[self.작업연도]
            else:
                return int(0)

        def 교직수당가산금1(self, 원로교사수당):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            생년월일 = rrn_to_datetime(self.교직원["주민번호"])
            작업월일 = (working_month)
            만나이 = 작업월일.year - 생년월일.year - ((작업월일.month, 작업월일.day) < (생년월일.month, 생년월일.day))
            비공무원경력일수 = 0
            if "경력" in self.교직원.keys():
                for 비공무원경력 in filter(lambda x:x[2]=="비공무원", self.교직원["경력"]):
                    비공무원경력일수 += ((self.strptime(비공무원경력[5])-self.strptime(비공무원경력[4])).days+1)*비공무원경력[3]/100
            근무연한일수 = (self.strptime(self.교직원["현근무년수변경일"])+relativedelta(years=self.교직원['근무연한'][0],months=self.교직원['근무연한'][1],days=self.교직원['근무연한'][2])-self.strptime(self.교직원["현근무년수변경일"])).days
            추가근무연한일수 = (working_month-self.strptime(self.교직원["현근무년수변경일"])).days+1
            근무연수 = int((근무연한일수 - 비공무원경력일수 + 추가근무연한일수)//365)
            if 만나이 >= 55 and 근무연수 >= 30:
                if self.교직원["원로교사"] == 0:
                    employee_elder_candidate_text.insert(END, f"작업연월, {작업월일}, 원로교사수당 지급 추정 추가 대상자:, {self.교직원['성명']}\n")
                    print("작업연월", 작업월일, "원로교사수당 지급 추정 추가 대상자:", self.교직원["성명"])
                    return 원로교사수당[self.작업연도] #자동입력, 자동입력 아닐 시 int(0)로 수정
                elif self.교직원["원로교사"] == 1:
                    return 원로교사수당[self.작업연도]
            else:
                return int(0)

        def 교직수당가산금2(self, 부장교사수당):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            if re.search("부장", self.교직원["보직"]):
                return 부장교사수당[self.작업연도]
            else:
                return int(0)

        def 교직수당가산금4(self, 담임교사수당):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            if re.search("담임", self.교직원["보직"]):
                return 담임교사수당[self.작업연도]
            else:
                return int(0)

        def 교직수당가산금6(self, 보건교사수당):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            if self.교직원["가산정원"] == "보건":
                return 보건교사수당[self.작업연도]
            else:
                return int(0)

        def 교직수당가산금10(self, 상담교사수당):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            if self.교직원["가산정원"] == "상담":
                return 상담교사수당[self.작업연도]
            else:
                return int(0)

        def 가족수당(self, 가족수당):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            try:
                self.교직원["가족사항"]
            except KeyError:
                return 0
            배우자수당 = len([가족 for 가족 in self.교직원["가족사항"] if 가족["가족관계"] == "배우자"])*가족수당[self.작업연도]["배우자"]
            직계존속수당 = 0
            for 가족 in [가족 for 가족 in self.교직원["가족사항"] if 가족["가족관계"] == "직계존속"]:
                생년월일 = rrn_to_datetime(가족["주민등록번호"])
                작업월일 = (working_month)
                만나이 = 작업월일.year - 생년월일.year - ((작업월일.month, 작업월일.day) < (생년월일.month, 생년월일.day))
                if (int(가족["주민등록번호"].split("-")[1][0]) in [1,3] and 만나이 >= 60) or (int(가족["주민등록번호"].split("-")[1][0]) in [2,4] and 만나이 >= 55):
                    직계존속수당 += 가족수당[self.작업연도]["직계존속"]
            직계비속수당 = 0
            직계비속생년월일 = []
            for 가족 in [가족 for 가족 in self.교직원["가족사항"] if 가족["가족관계"] == "직계비속"]:
                직계비속생년월일.append((가족["지급여부"], rrn_to_datetime(가족["주민등록번호"])))
            for n, 생년월일 in enumerate(sorted(직계비속생년월일, key= lambda x: x[1])):
                지급여부, 생년월일 = 생년월일
                작업월일 = (working_month)
                만나이 = 작업월일.year - 생년월일.year - ((작업월일.month, 작업월일.day) < (생년월일.month, 생년월일.day))
                if 만나이 < 20 and 지급여부 == 1:
                    직계비속수당 += 가족수당[self.작업연도]["직계비속"][n]
            return 배우자수당+직계존속수당+직계비속수당

        def 시간외근무수당정액분(self, 시간외근무수당):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            if self.교직원["직종"] == "행정직":
                return int(시간외근무수당[self.작업연도][self.교직원["직종"]][self.교직원["급"]]*10//10*10)
            elif self.교직원["보직"] == "교장":
                return int(0)
            elif self.교직원["보직"] == "교감":
                return int(시간외근무수당[self.작업연도][self.교직원["보직"]]*10//10*10)
            else:
                if self.교직원["직종"] == "기간제교원":
                    if int(self.교직원['승급년월일'].split("-")[1]) in [1,2]:
                        호봉 = self.교직원['호봉']+(working_month - datetime(int(self.교직원['승급년월일'].split("-")[0])-1,3,1)).days//365
                    else:
                        호봉 = self.교직원['호봉']+(working_month - datetime(int(self.교직원['승급년월일'].split("-")[0]),3,1)).days//365
                else:
                    호봉 = self.교직원['호봉']+(working_month - self.strptime(self.교직원['승급년월일'])).days//365
                return int([value for lower, upper, value in 시간외근무수당[self.작업연도]["교원"] if lower <= 호봉 and 호봉 <= upper][0]*10//10*10)
        def 학교운영수당(self, 학교운영수당):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            if self.교직원["직종"] == "행정직":
                return 학교운영수당[self.작업연도]
            else:
                return int(0)
        def 교원연구비(self):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return int(0)
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return int(0)
            if self.교직원["직종"] in ["교원", "기간제교원"]:
                if self.교직원["근무연한"][0] < 5:
                    return 75000
                return 60000
            return int(0)
        def 육아휴직수당(self, 육아휴직수당):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            values = []
            prolations = []
            working_month_totaldays = (working_month+relativedelta(months=1)-relativedelta(days=1)).day
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return values, prolations, working_month_totaldays
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return values, prolations, working_month_totaldays
            if "휴직" in self.교직원.keys():
                ongoing_parental_leaves = [[category, self.strptime(date_start), self.strptime(date_end), contigous, n] for n, [category, date_start, date_end, contigous] in enumerate(self.교직원["휴직"]) if re.search("육아휴직", category) and ((self.strptime(date_start)<=working_month+relativedelta(months=1)-relativedelta(days=1) and working_month<=self.strptime(date_start)) or (self.strptime(date_end)<=working_month+relativedelta(months=1)-relativedelta(days=1) and working_month<=self.strptime(date_end)) or (working_month<=self.strptime(date_end) and working_month>=self.strptime(date_start)))]
                for ongoing_parental_leave in ongoing_parental_leaves:
                    prolation = 0
                    prolation1, prolation2 = (working_month+relativedelta(months=1)-ongoing_parental_leave[1]).days, (ongoing_parental_leave[2]-working_month).days+1
                    if prolation1 <= working_month_totaldays and prolation2 <= working_month_totaldays:
                        prolation = prolation1 + prolation2 - working_month_totaldays
                    elif prolation1 <= working_month_totaldays:
                        prolation = prolation1
                    elif prolation2 <= working_month_totaldays:
                        prolation = prolation2
                    else:
                        prolation = working_month_totaldays
                    contigous = ongoing_parental_leave[3]
                    datelist = sorted([(self.strptime(date_start),self.strptime(date_end)) for category, date_start, date_end, _ in self.교직원["휴직"] if ongoing_parental_leave[0] == category and self.strptime(date_start)<=working_month+relativedelta(months=1)-relativedelta(days=1)], key=lambda x:x[0])
                    days = list(map(lambda datetime: working_month-datetime[0], [datelist[-1]]))
                    if len(datelist[:-1]) != 0:
                        days += map(lambda datetime: datetime[1]-datetime[0], datelist[:-1])
                    if len(days) > 1:
                        days = reduce(lambda x,y: x+y, days)
                    else:
                        days = days[0]
                    if days.days>=365:
                        value = int(0)
                    elif days.days<90: # 1~3개월
                        if contigous == 1:
                            if self.본봉()*육아휴직수당[self.작업연도][0][0]>2500000:
                                value = 2500000*0.85//10*10
                            elif self.본봉()*육아휴직수당[self.작업연도][0][0]<육아휴직수당[self.작업연도][0][2]:
                                value = 육아휴직수당[self.작업연도][0][2]*0.85//10*10
                            else:
                                value = self.본봉()*육아휴직수당[self.작업연도][0][0]*0.85//10*10
                        else:
                            if self.본봉()*육아휴직수당[self.작업연도][0][0]>육아휴직수당[self.작업연도][0][1]:
                                value = 육아휴직수당[self.작업연도][0][1]*0.85//10*10
                            elif self.본봉()*육아휴직수당[self.작업연도][0][0]<육아휴직수당[self.작업연도][0][2]:
                                value = 육아휴직수당[self.작업연도][0][2]*0.85//10*10
                            else:
                                value = self.본봉()*육아휴직수당[self.작업연도][0][0]*0.85//10*10
                    else: # 4~12개월
                        if self.본봉()*육아휴직수당[self.작업연도][1][0]>육아휴직수당[self.작업연도][1][1]:
                            value = 육아휴직수당[self.작업연도][1][1]*0.85//10*10
                        elif self.본봉()*육아휴직수당[self.작업연도][0][0]<육아휴직수당[self.작업연도][1][2]:
                            value = 육아휴직수당[self.작업연도][1][2]*0.85//10*10
                        else:
                            value = self.본봉()*육아휴직수당[self.작업연도][1][1]*0.85//10*10
                    values.append(value)
                    prolations.append(prolation)
                return values, prolations, working_month_totaldays
            else:
                return values, prolations, working_month_totaldays
        def 연가보상비(self):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            if "연가보상일수" in self.교직원.keys() and self.교직원["연가보상일수"] != "":
                if working_month.month == 7:
                    self.현재월 = 6
                    total = self.본봉()*0.86/30*5//10*10
                    self.현재월 = 7
                    return total
                elif working_month.month == 1:
                    제외월 = 0
                    if "휴직" in self.교직원.keys():
                        ongoing_leaves = [[self.strptime(date_start), self.strptime(date_end)] for _, date_start, date_end, _ in self.교직원["휴직"] if self.strptime(date_start) < datetime(int(작업연도), 1, 1) and datetime(int(작업연도), 12, 31) < self.strptime(date_end) or datetime(int(작업연도), 1, 1) <= self.strptime(date_start) and self.strptime(date_start) <= datetime(int(작업연도), 12, 31) or datetime(int(작업연도), 1, 1) <= self.strptime(date_end) and self.strptime(date_end) <= datetime(int(작업연도), 12, 31)]
                        for ongoing_leave in ongoing_leaves:
                            for i in range(1,13):
                                if (ongoing_leaves[0][0] - datetime(int(작업연도), i, 1)).days < 15 and (-15 < (ongoing_leaves[0][1] - datetime(int(작업연도), i, (datetime(int(작업연도), i, 1)+relativedelta(months=1)-relativedelta(days=1)).day)).days):
                                    제외월 += 1
                    연가보상일수 = int(int(self.교직원["연가보상일수"])*(12-제외월)/12)
                    self.현재월 = 12
                    total = self.본봉()*0.86/30*연가보상일수//10*10
                    self.현재월 = 6
                    partial = self.본봉()*0.86/30*5//10*10
                    self.현재월 = 13
                    return total - partial
                else:
                    return int(0)
            else:
                return int(0)
        # 정근수당도 수정 필요. 질병휴직 등은 근무한 월수로 안 침
        def 신분변동(self):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            percents = []
            prolations = []
            working_month_totaldays = (working_month+relativedelta(months=1)-relativedelta(days=1)).day
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return percents, prolations, working_month_totaldays
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return percents, prolations, working_month_totaldays
            if "휴직" in self.교직원.keys():
                ongoing_leaves = [[category, self.strptime(date_start), self.strptime(date_end)] for category, date_start, date_end, _ in self.교직원["휴직"] if not re.search("육아휴직", category) and ((self.strptime(date_start)<=working_month+relativedelta(months=1)-relativedelta(days=1) and working_month<=self.strptime(date_start)) or (self.strptime(date_end)<=working_month+relativedelta(months=1)-relativedelta(days=1) and working_month<=self.strptime(date_end)) or (working_month<=self.strptime(date_end) and working_month>=self.strptime(date_start)))]
                for ongoing_leave in ongoing_leaves:
                    prolation = 0
                    prolation1, prolation2 = (working_month+relativedelta(months=1)-ongoing_leave[1]).days, (ongoing_leave[2]-working_month).days+1
                    if prolation1 <= working_month_totaldays and prolation2 <= working_month_totaldays:
                        prolation = prolation1 + prolation2 - working_month_totaldays
                    elif prolation1 <= working_month_totaldays:
                        prolation = prolation1
                    elif prolation2 <= working_month_totaldays:
                        prolation = prolation2
                    else:
                        prolation = working_month_totaldays
                    days = (working_month - ongoing_leave[1]).days + 1
                    prolations.append(prolation)
                    if ongoing_leave[0] == "질병휴직":                        
                        if days <= 365:
                            percents.append(0.7)
                        elif days <= 365*2:
                            percents.append(0.5)
                        else:
                            percents.append(0)
                    else:
                        percents.append(0)
                return percents, prolations, working_month_totaldays
            else:
                return percents, prolations, working_month_totaldays

        def 징계(self):
            working_month = (datetime(int(작업연도), 1, 1)+relativedelta(months=self.현재월-1))
            percents = []
            prolations = []
            working_month_totaldays = (working_month+relativedelta(months=1)-relativedelta(days=1)).day
            if self.교직원['현부서임용일'] != "" and self.strptime(self.교직원['현부서임용일']) > working_month:
                return percents, prolations, working_month_totaldays
            if self.교직원['퇴직일'] != "" and self.strptime(self.교직원['퇴직일']) <= working_month:
                return percents, prolations, working_month_totaldays
            if "감봉시작일" in self.교직원.keys():
                date_start, date_end = self.strptime(self.교직원["감봉시작일"]), self.strptime(self.교직원["감봉종료일"])
                if ((date_start<=working_month+relativedelta(months=1)-relativedelta(days=1) and working_month<=date_start) or (date_end<=working_month+relativedelta(months=1)-relativedelta(days=1) and working_month<=date_end) or (working_month<=date_end and working_month>=date_start)):
                    ongoing_leave = [date_start, date_end]
                    prolation = 0
                    prolation1, prolation2 = (working_month+relativedelta(months=1)-ongoing_leave[0]).days, (ongoing_leave[1]-working_month).days+1
                    if prolation1 <= working_month_totaldays and prolation2 <= working_month_totaldays:
                        prolation = prolation1 + prolation2 - working_month_totaldays
                    elif prolation1 <= working_month_totaldays:
                        prolation = prolation1
                    elif prolation2 <= working_month_totaldays:
                        prolation = prolation2
                    else:
                        prolation = working_month_totaldays
                    days = (working_month - ongoing_leave[0]).days + 1
                    prolations.append(prolation)
                    if "감봉율" in self.교직원.keys():
                        percents.append(int(self.교직원["감봉율"])/100)
                    else:
                        percents.append(0)
                return percents, prolations, working_month_totaldays
            else:
                return percents, prolations, working_month_totaldays
    작업연도 = str(datetime.now().year)
    def set_working_year(event):
        global 작업연도
        작업연도 = employee_working_year_box.get()

    employee_working_year_label = Label(root, text="급여작업년도")
    employee_working_year_label.place(x=680, y=60+25+35)
    employee_working_year_box_var = [str(i) for i in range(datetime.now().year,2016,-1)]
    employee_working_year_box = ttk.Combobox(root, height=5, width=4, state='readonly', values=employee_working_year_box_var)
    employee_working_year_box.current(0)
    employee_working_year_box.place(x=700,y=85+25+35)
    employee_working_year_box.bind("<<ComboboxSelected>>", set_working_year)

    def create_salary(salary_creation):
        if "차기본봉표" in 급여생성.__dict__:
            del salary_creation.차기본봉표
        salary_creation.본봉표 = 본봉표(작업연도, 근속가봉표)
        try:
            salary_creation.차기본봉표 = 본봉표(str(int(작업연도)+1), 근속가봉표)
        except:
            pass
    salary_creation_btn = Button(root, text="작업연도 급여생성", command=lambda: create_salary(급여생성))
    salary_creation_btn.place(x=645,y=280+25+35-170)

    급여목록 = None
    def calculate_salary_table(employeelist):
        global 작업연도
        global 급여목록
        employee_elder_candidate_text.delete('1.0', END)
        급여목록 = [급여생성(employee, 작업연도, 명절년월일[작업연도]) for employee in employeelist]
        for 급여 in 급여목록:
            salarytables = []
            for 현재월 in range(3,15):
                급여.현재월 = 현재월
                if 현재월>12 and "차기본봉표" in 급여생성.__dict__:
                    급여.본봉표 = 급여.차기본봉표
                    급여.작업연도 = str(int(작업연도)+1)
                values, numers, denom = 급여.육아휴직수당(육아휴직수당)
                percents, prolations, _ = 급여.신분변동()
                percents2, prolations2, working_month_totaldays2 = 급여.징계()
                percents.append(1)
                prolations.append(denom-sum(numers)-sum(prolations))
                percents2.append(0)
                prolations2.append(denom-sum(numers)-sum(prolations2))
                value = sum(map(lambda value, numer : value*numer/denom, values, numers))
                salarytable = []
                salarytable.append(int(급여.본봉()*(denom-sum(numers))/denom*(1 if len(prolations)==1 else sum([prolation*percent for prolation, percent in zip(prolations, percents)])/(denom-sum(numers)))//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                salarytable.append(int(급여.정근수당()))
                salarytable.append(int(급여.정근수당가산금(가산금)*(denom-sum(numers))/denom*(1 if len(prolations)==1 else sum([prolation*percent for prolation, percent in zip(prolations, percents)])/(denom-sum(numers)))//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                salarytable.append(int(급여.정근수당추가가산금(추가가산금)*(denom-sum(numers))/denom*(1 if len(prolations)==1 else sum([prolation*percent for prolation, percent in zip(prolations, percents)])/(denom-sum(numers)))//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                if (denom-sum(numers)-sum(prolations[:-1]))>=15:
                    salarytable.append(int(급여.정액급식비(급식비[급여.작업연도])*(denom-sum(numers)-sum(prolations[:-1]))/denom//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                else:
                    salarytable.append(0)
                salarytable.append(int(급여.직급보조비(직급보조비)*(denom-sum(numers)-sum(prolations[:-1]))/denom//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                if sum(numers) == denom:
                    salarytable.append(0)
                else:
                    salarytable.append(급여.명절휴가비())
                salarytable.append(int(급여.관리업무수당()*(denom-sum(numers)-sum(prolations[:-1]))/denom//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                salarytable.append(int(급여.보전수당(보전수당)*(denom-sum(numers))/denom*(1 if len(prolations)==1 else sum([prolation*percent for prolation, percent in zip(prolations, percents)])/(denom-sum(numers)))//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                salarytable.append(int(급여.교직수당(교직수당)*(denom-sum(numers)-sum(prolations[:-1]))/denom//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                salarytable.append(int(급여.교직수당가산금1(원로교사수당)*(denom-sum(numers)-sum(prolations[:-1]))/denom//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                salarytable.append(int(급여.교직수당가산금2(부장교사수당)*(denom-sum(numers)-sum(prolations[:-1]))/denom//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                salarytable.append(int(급여.교직수당가산금4(담임교사수당)*(denom-sum(numers)-sum(prolations[:-1]))/denom//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                salarytable.append(int(급여.교직수당가산금6(보건교사수당)*(denom-sum(numers)-sum(prolations[:-1]))/denom//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                salarytable.append(int(급여.교직수당가산금10(상담교사수당)*(denom-sum(numers)-sum(prolations[:-1]))/denom//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                salarytable.append(int(급여.가족수당(가족수당)*(denom-sum(numers))/denom*(1 if len(prolations)==1 else sum([prolation*percent for prolation, percent in zip(prolations, percents)])/(denom-sum(numers)))//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                if (denom-sum(numers)-sum(prolations[:-1])-sum(prolations2[:-1]))>=15:
                    salarytable.append(급여.시간외근무수당정액분(시간외근무수당))
                else:
                    salarytable.append(int(급여.시간외근무수당정액분(시간외근무수당)*(denom-sum(numers)-sum(prolations[:-1])-sum(prolations2[:-1])))/15//10*10)
                salarytable.append(int(급여.학교운영수당(학교운영수당)*(denom-sum(numers)-sum(prolations[:-1]))/denom//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                salarytable.append(int(급여.교원연구비()*(denom-sum(numers))/denom*(1 if len(prolations)==1 else sum([prolation*percent for prolation, percent in zip(prolations, percents)])/(denom-sum(numers)))//10*10)*(sum(prolation*(1-percent) for prolation, percent in zip(prolations2, percents2))/denom)//10*10)
                salarytable.append(급여.연가보상비())
                salarytable.append(int(value//10*10))
                salarytables.append(salarytable)
            salarydf = pd.DataFrame(salarytables, columns=['본봉','정근수당','정근수당가산금','정근수당추가가산금','정액급식비','직급보조비','명절휴가비','관리업무수당','보전수당','교직수당','교직수당가산금1','교직수당가산금2','교직수당가산금4','교직수당가산금6','교직수당가산금10','가족수당','시간외근무수당정액분','학교운영수당','교원연구비','연가보상비','육아휴직수당'])
            salarydf = salarydf.set_axis([(datetime(int(작업연도), 3,1)+relativedelta(months=month)).date() for month in range(12)])
            급여.salary_table = salarydf

    salary_calculation_btn = Button(root, text="작업연도 급여계산", command=lambda: calculate_salary_table(employeelist))
    salary_calculation_btn.place(x=645,y=310+25+35-170)

    def save_salary_table(급여목록):
        wb = openpyxl.Workbook()
        ws = wb.active
        for 급여 in 급여목록:
            ws.append([급여.교직원["직종"], 급여.교직원["성명"], f"{급여.교직원['호봉']}호봉", 급여.교직원["급"], 급여.교직원["보직"], 급여.교직원["퇴직일"]])
            df = 급여.salary_table.transpose()
            df.insert(df.shape[1], "합계", df.sum(axis=1))
            for r in dataframe_to_rows(df, index=True, header=True):
                ws.append(r)
            ws.append([])
        wb.save(f"salarytable_작업연도{작업연도}_{datetime.now().strftime('%Y-%m-%d-%H%M%S')}.xlsx")
        showinfo("저장", "완료되었습니다!")
    salary_save_table_btn = Button(root, text="작업연도 급여저장", command=lambda: save_salary_table(급여목록))
    salary_save_table_btn.place(x=645,y=340+25+35-170)

    def save_salary_table_concat(급여목록):
        wb = openpyxl.Workbook()
        ws = wb.active
        if list(map(lambda category: [급여 for 급여 in 급여목록 if 급여.교직원["직종"] == category], ["행정직","교원","기간제교원"])) is None:
            return None
        for selected_category in list(map(lambda category: [급여 for 급여 in 급여목록 if 급여.교직원["직종"] == category], ["행정직","교원","기간제교원"])):
            if len(selected_category) == 0:
                continue
            ws.append([selected_category[0].교직원["직종"]])
            df = reduce(lambda x,y: x.add(y), [급여.salary_table for 급여 in selected_category])
            df = df.transpose()
            df.insert(df.shape[1], "합계", df.sum(axis=1))
            for r in dataframe_to_rows(df, index=True, header=True):
                ws.append(r)
            ws.append([])
        wb.save(f"salarytable_concat_작업연도{작업연도}_{datetime.now().strftime('%Y-%m-%d-%H%M%S')}.xlsx")
        showinfo("저장", "완료되었습니다!")

    salary_save_table_btn = Button(root, text="작업연도 급여합산", command=lambda: save_salary_table_concat(급여목록))
    salary_save_table_btn.place(x=645,y=370+25+35-170)
    ##############  Code   ###############

    def load_employeelist(root):
        subroot = Toplevel(root)
        subroot.title("New Window")
        subroot.geometry("300x80")
        Label(subroot, text="파일목록").pack()
        subroot_box_var = list(reversed(sorted([fname for fname in os.listdir(".") if re.search(".*\.pickle", fname)])))
        subroot_box = ttk.Combobox(subroot, width=35, state='readonly', values=subroot_box_var)
        subroot_box.current(0)
        subroot_box.pack()
        def read_employeelist():
            global employeelist_treeview
            global employeelist
            with open(subroot_box.get(), "rb") as f:
                employeelist = pickle.load(f)
            subroot.destroy()
            for i in employeelist_treeview.get_children():
                employeelist_treeview.delete(i)
            [employeelist_treeview.insert("",END,values=(employee["id"],employee["성명"])) for employee in employeelist]
        subroot_btn = Button(subroot, text="불러오기", command=lambda:read_employeelist())
        subroot_btn.pack()

    salary_load_employeelist_btn = Button(root, text="인사기록 불러오기", command=lambda: load_employeelist(root))
    salary_load_employeelist_btn.place(x=645,y=400+25+35-170)
    def save_employeelist():
        with open(f"employeelist_{datetime.now().strftime('%Y-%m-%d-%H%M%S')}.pickle", "wb") as f:
            pickle.dump(employeelist, f)
    salary_save_employeelist_btn = Button(root, text="인사기록 저장하기", command=lambda: save_employeelist())
    salary_save_employeelist_btn.place(x=645,y=430+25+35-170)

    def save_salary_table_for_summaries(급여목록):
        wb = openpyxl.Workbook()
        ws = wb.active
        for 급여 in 급여목록:
            temp = 급여.salary_table["명절휴가비"].tolist()
            temp = [i for i in temp if i != 0]                
            df = 급여.salary_table               
            df = df.sum(axis=0)
            if 급여.교직원["퇴직일"] != "" and datetime(*map(int, 급여.교직원["퇴직일"].split("-"))) < datetime(int(작업연도), 3, 1):
                continue
            ws.append([급여.교직원["성명"],
                       rrn_to_datetime(급여.교직원["주민번호"]).date(),
                       급여.교직원["직종"],
                       '',
                       '',
                       급여.교직원["현부서임용일"],
                       "",
                       (relativedelta(years = 급여.교직원["근무연한"][0], months = 급여.교직원["근무연한"][1], days = 급여.교직원["근무연한"][2]) + datetime(int(작업연도), 3, 1) - datetime(*map(int, 급여.교직원["승급년월일"].split("-")))).days//365,
                       급여.교직원['호봉']+(datetime(int(작업연도), 3, 1) - datetime(*map(int, 급여.교직원["승급년월일"].split("-")))).days//365,
                       "-".join(급여.교직원["승급년월일"].split("-")[1:]),
                       df["본봉"],
                       df["정근수당"],
                       df["정근수당가산금"],
                       df["정근수당추가가산금"],
                       df["정액급식비"],
                       df["교직수당"],
                       df["교직수당가산금1"],
                       df["교직수당가산금2"],
                       "",
                       df["교직수당가산금4"],
                       "",
                       df["교직수당가산금6"],
                       "",
                       "",
                       "",
                       df["교직수당가산금10"],
                       df["보전수당"],
                       df["교원연구비"],
                       0 if sum(temp) == 0 else temp[0],
                       0 if sum(temp) == 0 else temp[1],
                       df["명절휴가비"],
                       "",
                       "",
                       "",
                       df["시간외근무수당정액분"] if 급여.교직원["직종"]=="행정직" else df["시간외근무수당정액분"]*11/12//10*10,
                       df["시간외근무수당정액분"]/12//10*10,
                       df["관리업무수당"],
                       df["육아휴직수당"],
                       df["직급보조비"],
                       df["연가보상비"],
                       "",
                       "",
                       df["학교운영수당"],
                       "",
                       "",
                       ""
                       ])
        wb.save(f"salarytable_작업연도{작업연도}_총괄표_{datetime.now().strftime('%Y-%m-%d-%H%M%S')}.xlsx")
        showinfo("저장", "완료되었습니다!")
    save_salary_table_for_summaries_btn = Button(root, text="재정결함 총괄표저장", command=lambda: save_salary_table_for_summaries(급여목록))
    save_salary_table_for_summaries_btn.place(x=645,y=460+25+35-170)

    # root.config(menu=menu)
    root.mainloop()
